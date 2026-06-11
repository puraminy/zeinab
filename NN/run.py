from matplotlib import pyplot as plt
import torch
import torch.nn as nn
import torch.optim as optim
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.model_selection import KFold, RandomizedSearchCV, train_test_split
from sklearn.ensemble import (
    RandomForestRegressor,
    ExtraTreesRegressor,
    GradientBoostingRegressor,
    AdaBoostRegressor,
    HistGradientBoostingRegressor,
)
from sklearn.neighbors import KNeighborsRegressor
from sklearn.svm import SVR
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet
from sklearn.tree import DecisionTreeRegressor
from sklearn.multioutput import MultiOutputRegressor
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import ConstantKernel, RBF, WhiteKernel
import pandas as pd
import numpy as np
import re
from sklearn.metrics import make_scorer, r2_score
from sklearn.inspection import permutation_importance
from sklearn.feature_selection import mutual_info_regression
from tabulate import tabulate
import os
from latex import *
from plot import *
from models import *
from read_data import (
    read_prep_data,
    sync_prep_data_with_dataset,
    prep_data_exists,
    read_prep_metadata,
    save_data,
    print_path_debug,
    ensure_dataset_csv_exists,
)
from refinery_variables import (
    CONTROL_VARIABLES,
    EARLY_VARIABLES,
    TARGET_VARIABLES,
    filter_allowed_model_inputs,
    find_leakage_columns,
    leakage_pattern_matches,
    refinery_variable_group_metadata,
    remove_name_based_leakage_inputs,
    validate_model_inputs,
)
from recommendation_engine import (
    DEFAULT_RECOMMENDED_CONTROLS,
    RecommendationError,
    recommend_operating_conditions,
)
import inspect
import models
import json
import copy
import math
from datetime import datetime
from collections import OrderedDict

from openpyxl.styles import Font, PatternFill, Alignment

try:
    from xgboost import XGBRegressor
except ImportError:
    XGBRegressor = None

try:
    from lightgbm import LGBMRegressor
except ImportError:
    LGBMRegressor = None

try:
    from catboost import CatBoostRegressor
except ImportError:
    CatBoostRegressor = None

ANSI_RESET = "\033[0m"
ANSI_GREEN = "\033[92m"
ANSI_BLUE = "\033[94m"
SAVED_RUNS_DIR = "saved_runs"
CHECKPOINTS_DIR = "checkpoints"
MODULE_DIR = os.path.dirname(os.path.abspath(__file__))
LEAKAGE_REPORT_PATH = os.path.join(MODULE_DIR, "reports", "leakage_report.xlsx")
FEATURE_IMPORTANCE_REPORTS_DIR = os.path.join(MODULE_DIR, "reports")
MODEL_COMPARISON_REPORT_PATH = os.path.join(MODULE_DIR, "reports", "model_comparison.xlsx")
OPTIMIZATION_REPORT_PATH = os.path.join(MODULE_DIR, "reports", "optimization_report.xlsx")
OPERATOR_REPORT_PATH = os.path.join(MODULE_DIR, "reports", "operator_report.xlsx")


FEATURE_IMPORTANCE_TARGETS = (
    "white_total_points",
    "white_solution_color",
    "white_apparent_color",
    "white_ash",
    "white_moisture",
    "white_invert",
)




PRESCRIPTIVE_OPTIMIZATION_VARIABLES = (
    "lime_milk_baume",
    "lime_alkalinity",
    "co2_percent",
    "carbonated_pH",
    "sulphited_pH",
    "sulphited_brix",
    "standard_liquor_pH",
    "standard_liquor_brix",
)

OPERATOR_RECOMMENDATION_VARIABLES = DEFAULT_RECOMMENDED_CONTROLS

PRESCRIPTIVE_OPTIMIZATION_OBJECTIVES = (
    "white_total_points",
    "white_solution_color",
    "white_apparent_color",
    "white_ash",
)


FUTURE_QUALITY_INPUT_CANDIDATES = [
    "lime_milk_baume",
    "filtercake_moisture",
    "filtercake_sugar",
    "sweetwater_brix",
    "sulphited_pH",
    "sulphited_brix",
    "sulphited_color",
    "standard_liquor_pH",
    "standard_liquor_brix",
    "standard_liquor_color",
]




def coerce_refinery_numeric_series(series):
    """Convert numeric-like refinery strings, including ranges such as 10-12, to floats."""
    def _convert(value):
        if pd.isna(value):
            return np.nan
        if isinstance(value, (int, float, np.integer, np.floating)):
            return float(value)
        text = str(value).strip()
        if not text:
            return np.nan
        normalized = (
            text.replace("−", "-")
            .replace("–", "-")
            .replace("—", "-")
            .replace(",", "")
        )
        direct_value = pd.to_numeric(normalized, errors="coerce")
        if pd.notna(direct_value):
            return float(direct_value)
        range_match = re.fullmatch(
            r"\s*([+-]?\d+(?:\.\d+)?)\s*-\s*([+-]?\d+(?:\.\d+)?)\s*",
            normalized,
        )
        if range_match:
            low, high = map(float, range_match.groups())
            return (low + high) / 2.0
        return np.nan

    return series.map(_convert).astype(float)


def coerce_refinery_numeric_frame(dataframe):
    """Convert every column in a dataframe with the refinery numeric parser."""
    return dataframe.apply(coerce_refinery_numeric_series, axis=0)

def color_text(text, color):
    return f"{color}{text}{ANSI_RESET}"


def print_divider(char="=", width=68):
    print(char * width)


def print_numbered_feature_list(title, features, color):
    print_divider("=")
    print(title)
    print_divider("-")
    for idx, feature_name in enumerate(features, start=1):
        print(color_text(f"{idx:>2}. {feature_name}", color))
    print_divider("=")


def save_leakage_report(removed_columns, output_features=None, context="pre_training", report_path=LEAKAGE_REPORT_PATH):
    """Write the target-leakage audit workbook showing exact X removals."""
    output_features = list(output_features or [])
    removed_rows = [
        {
            "context": context,
            "column": column,
            "matched_patterns": ", ".join(leakage_pattern_matches(column)),
            "action": "removed_from_X",
            "reason": "Column name contains an automatic target-leakage marker.",
        }
        for column in removed_columns
    ]
    kept_target_rows = [
        {
            "context": context,
            "column": column,
            "matched_patterns": ", ".join(leakage_pattern_matches(column)),
            "action": "kept_as_final_target",
            "reason": "Selected final targets are not removed from y.",
        }
        for column in output_features
        if leakage_pattern_matches(column)
    ]
    summary_rows = [
        {
            "context": context,
            "created_at_utc": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "removed_from_X_count": len(removed_rows),
            "kept_final_target_count": len(kept_target_rows),
            "report_path": report_path,
        }
    ]

    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        pd.DataFrame(removed_rows, columns=["context", "column", "matched_patterns", "action", "reason"]).to_excel(
            writer, sheet_name="removed_from_X", index=False
        )
        pd.DataFrame(kept_target_rows, columns=["context", "column", "matched_patterns", "action", "reason"]).to_excel(
            writer, sheet_name="kept_targets", index=False
        )
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="summary", index=False)
    print(f"Saved target-leakage report: {report_path}")


def remove_leakage_inputs_for_training(input_features, output_features=None, context="pre_training"):
    """Print, report, and remove automatic name-based leakage columns from X."""
    cleaned_inputs, removed_columns = remove_name_based_leakage_inputs(
        list(input_features),
        output_features=output_features,
    )
    if removed_columns:
        print("Target leakage columns detected in X candidates:")
        for column in removed_columns:
            patterns = ", ".join(leakage_pattern_matches(column))
            print(f" - {column} (matched: {patterns})")
        print("Removed leakage columns from X: " + ", ".join(removed_columns))
    else:
        print("No automatic name-based target leakage columns detected in X candidates.")

    kept_targets = [column for column in (output_features or []) if leakage_pattern_matches(column)]
    if kept_targets:
        print("Final target column(s) kept in y (not removed): " + ", ".join(kept_targets))

    save_leakage_report(removed_columns, output_features=output_features, context=context)
    return cleaned_inputs, removed_columns





def _format_missing_percentage(value):
    """Format a missing-value percentage consistently for console reports."""
    return f"{value:.2f}%"


def _print_missing_value_report(X_train, X_test, train_missing_counts, test_missing_counts):
    """Print counts and percentages for columns that contain missing feature values."""
    all_missing_columns = [
        column
        for column in X_train.columns
        if train_missing_counts.get(column, 0) > 0 or test_missing_counts.get(column, 0) > 0
    ]
    if not all_missing_columns:
        print("Missing-value report: no NaN values detected in selected input columns.")
        return

    print("\n================= Missing-Value Report =================")
    print("Columns containing NaN values:")
    for column in all_missing_columns:
        train_count = int(train_missing_counts.get(column, 0))
        test_count = int(test_missing_counts.get(column, 0))
        train_pct = (train_count / max(len(X_train), 1)) * 100.0
        test_pct = (test_count / max(len(X_test), 1)) * 100.0
        print(
            f" - {column}: "
            f"train={train_count}/{len(X_train)} ({_format_missing_percentage(train_pct)}), "
            f"test={test_count}/{len(X_test)} ({_format_missing_percentage(test_pct)})"
        )
        if train_pct > 20.0 or test_pct > 20.0:
            print(
                f"   WARNING: missing percentage exceeds 20% "
                f"(train={_format_missing_percentage(train_pct)}, "
                f"test={_format_missing_percentage(test_pct)})."
            )
    print("========================================================\n")


def apply_missing_value_pipeline(X_train, X_test, missing_drop_threshold=50.0, verbose=True):
    """Drop high-missing feature columns and median-impute remaining values.

    Missingness thresholds and median imputation are fitted from training data only;
    the learned column set and medians are then applied to test data.
    """
    X_train_clean = X_train.copy().replace([np.inf, -np.inf], np.nan)
    X_test_clean = X_test.copy().replace([np.inf, -np.inf], np.nan)
    X_train_clean = X_train_clean.apply(pd.to_numeric, errors="coerce")
    X_test_clean = X_test_clean.apply(pd.to_numeric, errors="coerce")

    train_missing_counts = X_train_clean.isna().sum()
    test_missing_counts = X_test_clean.isna().sum()
    if verbose:
        _print_missing_value_report(X_train_clean, X_test_clean, train_missing_counts, test_missing_counts)

    train_missing_percentages = train_missing_counts / max(len(X_train_clean), 1) * 100.0
    columns_to_drop = train_missing_percentages[train_missing_percentages > missing_drop_threshold].index.tolist()
    if columns_to_drop:
        if verbose:
            print(
                f"Dropping {len(columns_to_drop)} input column(s) with >{missing_drop_threshold:.0f}% "
                "missing values in training data: "
                + ", ".join(columns_to_drop)
            )
        X_train_clean = X_train_clean.drop(columns=columns_to_drop)
        X_test_clean = X_test_clean.drop(columns=columns_to_drop, errors="ignore")

    if X_train_clean.empty:
        raise ValueError("No input columns remain after dropping columns with >50% missing training values.")

    imputer = SimpleImputer(strategy="median")
    X_train_imputed = pd.DataFrame(
        imputer.fit_transform(X_train_clean),
        columns=X_train_clean.columns,
        index=X_train_clean.index,
    )
    X_test_imputed = pd.DataFrame(
        imputer.transform(X_test_clean.reindex(columns=X_train_clean.columns)),
        columns=X_train_clean.columns,
        index=X_test_clean.index,
    )
    X_train_imputed = X_train_imputed.astype(float)
    X_test_imputed = X_test_imputed.astype(float)
    return X_train_imputed, X_test_imputed, columns_to_drop


def _safe_excel_sheet_name(name):
    """Return an Excel-compatible sheet name."""
    cleaned = re.sub(r"[\\/*?:\[\]]+", "_", str(name)).strip() or "Sheet"
    return cleaned[:31]


def _safe_report_filename_part(name):
    """Return a filesystem-safe filename part for a target variable."""
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(name)).strip("_") or "target"


def _coerce_feature_importance_source(raw_df, target, seed=None, test_size=0.25):
    """Prepare train/holdout feature matrices without fitting on holdout rows.

    Numeric medians, categorical levels, and constant-column filtering are learned
    from the feature-importance training split only, then applied to the
    permutation holdout split.  This keeps preprocessing consistent with the
    model-evaluation path and avoids letting holdout rows influence imputation
    or encoding.
    """
    if seed is None:
        seed = model_seed
    if target not in raw_df.columns:
        return None, None, None, None, None

    working_df = raw_df.copy().replace([np.inf, -np.inf], np.nan)
    y = coerce_refinery_numeric_series(working_df[target])
    feature_columns = [
        column for column in working_df.columns
        if column != target and not leakage_pattern_matches(column)
    ]
    if not feature_columns:
        return None, None, None, None, None

    X_raw = working_df[feature_columns].copy()
    valid_rows = y.notna()
    X_raw = X_raw.loc[valid_rows]
    y = y.loc[valid_rows].astype(float)
    if len(y) < 3:
        return None, None, None, None, None

    if len(y) >= 8:
        X_train_raw, X_perm_raw, y_train, y_perm = train_test_split(
            X_raw,
            y,
            test_size=test_size,
            random_state=seed,
        )
    else:
        # Very small report-only datasets cannot spare a holdout; no separate
        # test preprocessing is fitted because train and permutation data are
        # intentionally identical in this fallback.
        X_train_raw, X_perm_raw, y_train, y_perm = X_raw, X_raw, y, y

    encoded_train_parts = []
    encoded_perm_parts = []
    encoded_to_original = {}

    for column in feature_columns:
        train_series = X_train_raw[column]
        train_non_null = train_series.notna().sum()
        if train_non_null == 0:
            continue

        train_numeric = coerce_refinery_numeric_series(train_series)
        train_numeric_ratio = train_numeric.notna().sum() / max(train_non_null, 1)
        if pd.api.types.is_numeric_dtype(train_series) or train_numeric_ratio >= 0.8:
            fill_value = train_numeric.median()
            if pd.isna(fill_value):
                continue
            perm_numeric = coerce_refinery_numeric_series(X_perm_raw[column])
            encoded_train_parts.append(train_numeric.fillna(fill_value).astype(float).to_frame(column))
            encoded_perm_parts.append(perm_numeric.fillna(fill_value).astype(float).to_frame(column))
            encoded_to_original[column] = column
        else:
            train_categorical = train_series.astype("object").where(train_series.notna(), "__missing__").astype(str)
            train_dummies = pd.get_dummies(
                train_categorical,
                prefix=column,
                prefix_sep="__category__",
                dummy_na=False,
                dtype=float,
            )
            if train_dummies.empty:
                continue
            perm_categorical = X_perm_raw[column].astype("object").where(
                X_perm_raw[column].notna(), "__missing__"
            ).astype(str)
            perm_dummies = pd.get_dummies(
                perm_categorical,
                prefix=column,
                prefix_sep="__category__",
                dummy_na=False,
                dtype=float,
            ).reindex(columns=train_dummies.columns, fill_value=0.0)
            encoded_train_parts.append(train_dummies)
            encoded_perm_parts.append(perm_dummies)
            for encoded_column in train_dummies.columns:
                encoded_to_original[encoded_column] = column

    if not encoded_train_parts:
        return None, None, None, None, None

    X_train_encoded = pd.concat(encoded_train_parts, axis=1)
    X_perm_encoded = pd.concat(encoded_perm_parts, axis=1).reindex(columns=X_train_encoded.columns, fill_value=0.0)

    constant_columns = [
        column for column in X_train_encoded.columns
        if X_train_encoded[column].nunique(dropna=False) <= 1
    ]
    if constant_columns:
        X_train_encoded = X_train_encoded.drop(columns=constant_columns)
        X_perm_encoded = X_perm_encoded.drop(columns=constant_columns, errors="ignore")
        for column in constant_columns:
            encoded_to_original.pop(column, None)
    if X_train_encoded.empty:
        return None, None, None, None, None

    return X_train_encoded, X_perm_encoded, y_train, y_perm, encoded_to_original


def _aggregate_encoded_importance(encoded_values, encoded_columns, encoded_to_original):
    """Aggregate one-hot encoded importances back to source variable names."""
    totals = {}
    for encoded_column, value in zip(encoded_columns, encoded_values):
        original_column = encoded_to_original.get(encoded_column, encoded_column)
        totals[original_column] = totals.get(original_column, 0.0) + float(value)
    return pd.DataFrame(
        [{"Variable": variable, "Importance": importance} for variable, importance in totals.items()]
    ).sort_values(by="Importance", ascending=False, kind="mergesort").reset_index(drop=True)


def _rank_importance_table(table, importance_column="Importance"):
    """Add a dense rank to a variable-importance table."""
    ranked = table.copy()
    ranked["Rank"] = ranked[importance_column].rank(method="dense", ascending=False).astype(int)
    return ranked[["Rank", "Variable", importance_column]]


def _build_feature_importance_tables(dataframe, target, seed=None):
    """Calculate tree, permutation, and mutual-information importance tables."""
    if seed is None:
        seed = model_seed
    X_train_fi, X_perm, y_train_fi, y_perm, encoded_to_original = _coerce_feature_importance_source(
        dataframe,
        target,
        seed=seed,
    )
    if X_train_fi is None:
        return None

    rf_model = RandomForestRegressor(
        n_estimators=500,
        random_state=seed,
        n_jobs=-1,
    )
    rf_model.fit(X_train_fi, y_train_fi)

    et_model = ExtraTreesRegressor(
        n_estimators=500,
        random_state=seed,
        n_jobs=-1,
    )
    et_model.fit(X_train_fi, y_train_fi)

    rf_table = _rank_importance_table(
        _aggregate_encoded_importance(rf_model.feature_importances_, X_train_fi.columns, encoded_to_original)
    )
    et_table = _rank_importance_table(
        _aggregate_encoded_importance(et_model.feature_importances_, X_train_fi.columns, encoded_to_original)
    )

    permutation = permutation_importance(
        rf_model,
        X_perm,
        y_perm,
        n_repeats=10,
        random_state=seed,
        n_jobs=-1,
        scoring="r2",
    )
    permutation_table = _aggregate_encoded_importance(
        permutation.importances_mean,
        X_train_fi.columns,
        encoded_to_original,
    )
    permutation_std_table = _aggregate_encoded_importance(
        permutation.importances_std,
        X_train_fi.columns,
        encoded_to_original,
    ).rename(columns={"Importance": "Permutation Std"})
    permutation_table = permutation_table.merge(permutation_std_table, on="Variable", how="left")
    permutation_table = _rank_importance_table(permutation_table)
    permutation_table = permutation_table.merge(
        permutation_std_table,
        on="Variable",
        how="left",
    )[["Rank", "Variable", "Importance", "Permutation Std"]]

    mutual_information = mutual_info_regression(X_train_fi, y_train_fi, random_state=seed)
    mi_table = _rank_importance_table(
        _aggregate_encoded_importance(mutual_information, X_train_fi.columns, encoded_to_original)
    )

    return {
        "Random Forest Importance": rf_table,
        "Extra Trees Importance": et_table,
        "Permutation Importance": permutation_table,
        "Mutual Information": mi_table,
        "metadata": pd.DataFrame([
            {
                "target": target,
                "rows_used": int(len(y_train_fi.index.union(y_perm.index))),
                "training_rows_used": int(len(y_train_fi)),
                "permutation_rows_used": int(len(y_perm)),
                "source_variables_ranked": int(len(set(encoded_to_original.values()))),
                "encoded_feature_count": int(X_train_fi.shape[1]),
                "preprocessing_scope": "fit_on_training_split_only",
                "created_at_utc": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            }
        ]),
    }


def _combine_feature_importance_ranks(tables):
    """Create one combined table with scores/ranks from all requested methods."""
    method_names = [
        "Random Forest Importance",
        "Extra Trees Importance",
        "Permutation Importance",
        "Mutual Information",
    ]
    combined = None
    for method_name in method_names:
        method_table = tables[method_name][["Variable", "Importance", "Rank"]].rename(
            columns={
                "Importance": f"{method_name} Score",
                "Rank": f"{method_name} Rank",
            }
        )
        combined = method_table if combined is None else combined.merge(method_table, on="Variable", how="outer")

    rank_columns = [f"{method_name} Rank" for method_name in method_names]
    combined["Average Rank"] = combined[rank_columns].mean(axis=1)
    combined = combined.sort_values(by=["Average Rank", "Variable"], ascending=[True, True]).reset_index(drop=True)
    combined.insert(0, "Overall Rank", np.arange(1, len(combined) + 1))
    return combined


def _load_feature_importance_source_dataframe(dataset_path, X_train, X_test, y_train, y_test):
    """Prefer stclean.csv for fixed-target reports, then fall back to prep_data."""
    candidate_paths = [
        dataset_path,
        os.path.join("convert", "stclean.csv"),
        os.path.join("convert", "sugar_all.csv"),
    ]
    for candidate_path in candidate_paths:
        try:
            resolved_path = ensure_dataset_csv_exists(candidate_path)
            if os.path.isfile(resolved_path):
                print(f"Feature-importance source dataset: {resolved_path}")
                return pd.read_csv(resolved_path)
        except (FileNotFoundError, ImportError, ValueError, OSError) as err:
            print(f"Feature-importance source skipped ({candidate_path}): {err}")

    print("Feature-importance source fallback: combined prep_data train/test matrices.")
    train_df = pd.concat([X_train.reset_index(drop=True), y_train.reset_index(drop=True)], axis=1)
    test_df = pd.concat([X_test.reset_index(drop=True), y_test.reset_index(drop=True)], axis=1)
    return pd.concat([train_df, test_df], axis=0, ignore_index=True)


def generate_feature_importance_reports(dataframe, targets=FEATURE_IMPORTANCE_TARGETS, reports_dir=FEATURE_IMPORTANCE_REPORTS_DIR):
    """Generate one ranked Excel feature-importance workbook per requested target."""
    os.makedirs(reports_dir, exist_ok=True)
    generated_reports = []
    print("\n================= Feature Importance Reports =================")
    for target in targets:
        if target not in dataframe.columns:
            print(f"Skipping {target}: target column was not found in the feature-importance dataset.")
            continue

        print(f"Calculating feature importance for target: {target}")
        tables = _build_feature_importance_tables(dataframe, target)
        if tables is None:
            print(f"Skipping {target}: not enough valid numeric target rows/features for importance analysis.")
            continue

        combined = _combine_feature_importance_ranks(tables)
        top_20 = combined.head(20)
        report_path = os.path.join(
            reports_dir,
            f"feature_importance_{_safe_report_filename_part(target)}.xlsx",
        )
        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
            combined.to_excel(writer, sheet_name="All Methods", index=False)
            top_20.to_excel(writer, sheet_name="Top 20", index=False)
            for sheet_name in [
                "Random Forest Importance",
                "Extra Trees Importance",
                "Permutation Importance",
                "Mutual Information",
            ]:
                tables[sheet_name].to_excel(
                    writer,
                    sheet_name=_safe_excel_sheet_name(sheet_name),
                    index=False,
                )
            tables["metadata"].to_excel(writer, sheet_name="Metadata", index=False)

        generated_reports.append(report_path)
        print(f"Saved feature-importance report: {report_path}")
        print(f"Top 20 variables for {target} (ranked by average rank across methods):")
        display_columns = ["Overall Rank", "Variable", "Average Rank"]
        print(tabulate(top_20[display_columns], headers="keys", tablefmt="github", showindex=False))
    print("==============================================================\n")
    return generated_reports



def _normal_cdf(values):
    """Vectorized standard-normal CDF without requiring scipy."""
    arr = np.asarray(values, dtype=float)
    return 0.5 * (1.0 + np.vectorize(math.erf)(arr / math.sqrt(2.0)))


def _normal_pdf(values):
    """Vectorized standard-normal PDF without requiring scipy."""
    arr = np.asarray(values, dtype=float)
    return np.exp(-0.5 * arr ** 2) / math.sqrt(2.0 * math.pi)


def _prepare_prescriptive_optimization_data(dataframe, control_variables, objectives):
    """Return numeric, complete historical rows for the requested optimization task."""
    requested_columns = list(control_variables) + list(objectives)
    missing_columns = [column for column in requested_columns if column not in dataframe.columns]
    if missing_columns:
        return None, missing_columns

    numeric_df = coerce_refinery_numeric_frame(dataframe[requested_columns]).replace([np.inf, -np.inf], np.nan)
    complete_df = numeric_df.dropna(axis=0, how="any").reset_index(drop=True)
    if complete_df.empty:
        return None, []
    return complete_df, []


def _objective_normalization_parameters(objective_df):
    """Build robust min/max parameters so differently scaled objectives can be summed."""
    objective_min = objective_df.min(axis=0)
    objective_max = objective_df.max(axis=0)
    objective_range = (objective_max - objective_min).replace(0, 1.0)
    return objective_min, objective_range


def _scalar_quality_objective(objective_values, objective_min, objective_range):
    """Lower is better: mean min/max-normalized white sugar quality penalties."""
    values = pd.DataFrame(objective_values, columns=list(objective_min.index))
    normalized = (values - objective_min) / objective_range
    return normalized.mean(axis=1).to_numpy(dtype=float)


def _predict_scalar_quality(candidate_values, model, objective_min, objective_range):
    """Predict the scalar minimization objective for one or many candidate rows."""
    candidate_array = np.asarray(candidate_values, dtype=float)
    if candidate_array.ndim == 1:
        candidate_array = candidate_array.reshape(1, -1)
    predicted_objectives = model.predict(candidate_array)
    return _scalar_quality_objective(predicted_objectives, objective_min, objective_range)


def _clip_to_bounds(candidate, lower_bounds, upper_bounds):
    return np.minimum(np.maximum(np.asarray(candidate, dtype=float), lower_bounds), upper_bounds)


def _bayesian_optimize_controls(X, scalar_y, lower_bounds, upper_bounds, seed=123, iterations=3, candidates_per_iter=400):
    """Gaussian-process Bayesian optimization with Expected Improvement acquisition."""
    rng = np.random.default_rng(seed)
    X_df = dataframe_from_tabular(X, "X")
    x_scaler = StandardScaler()
    X_scaled_df = fit_transform_dataframe(x_scaler, X_df)
    lower_df = pd.DataFrame([lower_bounds], columns=X_df.columns)
    upper_df = pd.DataFrame([upper_bounds], columns=X_df.columns)
    lower_scaled = transform_dataframe(x_scaler, lower_df).to_numpy().reshape(-1)
    upper_scaled = transform_dataframe(x_scaler, upper_df).to_numpy().reshape(-1)
    X_scaled = X_scaled_df.to_numpy()
    scaled_low = np.minimum(lower_scaled, upper_scaled)
    scaled_high = np.maximum(lower_scaled, upper_scaled)

    kernel = ConstantKernel(1.0, (1e-3, 1e3)) * RBF(length_scale=np.ones(X.shape[1]), length_scale_bounds=(1e-2, 1e2)) + WhiteKernel(noise_level=1e-5, noise_level_bounds=(1e-8, 1e-1))
    optimizer_X = X_scaled.copy()
    optimizer_y = np.asarray(scalar_y, dtype=float).copy()
    best_scaled = optimizer_X[int(np.argmin(optimizer_y))].copy()

    for _ in range(int(iterations)):
        gp = GaussianProcessRegressor(
            kernel=kernel,
            normalize_y=True,
            random_state=seed,
            optimizer=None,
            n_restarts_optimizer=0,
        )
        gp.fit(optimizer_X, optimizer_y)
        random_candidates = rng.uniform(scaled_low, scaled_high, size=(int(candidates_per_iter), X.shape[1]))
        mean_pred, std_pred = gp.predict(random_candidates, return_std=True)
        std_pred = np.maximum(std_pred, 1e-9)
        improvement = float(np.min(optimizer_y)) - mean_pred
        z_score = improvement / std_pred
        expected_improvement = improvement * _normal_cdf(z_score) + std_pred * _normal_pdf(z_score)
        best_candidate_index = int(np.argmax(expected_improvement))
        chosen_scaled = random_candidates[best_candidate_index]
        chosen_score = float(mean_pred[best_candidate_index])
        optimizer_X = np.vstack([optimizer_X, chosen_scaled])
        optimizer_y = np.append(optimizer_y, chosen_score)
        if chosen_score < float(np.min(optimizer_y[:-1])):
            best_scaled = chosen_scaled.copy()

    best_scaled_df = pd.DataFrame([best_scaled], columns=X_df.columns)
    recommended = inverse_transform_dataframe(x_scaler, best_scaled_df).to_numpy().reshape(-1)
    return _clip_to_bounds(recommended, lower_bounds, upper_bounds)


def _differential_evolution_optimize_controls(
    objective_function,
    lower_bounds,
    upper_bounds,
    seed=123,
    population_size=5,
    generations=12,
    mutation=0.7,
    crossover=0.8,
):
    """Small self-contained Differential Evolution implementation for bounded controls."""
    rng = np.random.default_rng(seed)
    dimensions = len(lower_bounds)
    population_count = max(int(population_size) * dimensions, 24)
    population = rng.uniform(lower_bounds, upper_bounds, size=(population_count, dimensions))
    scores = np.asarray([objective_function(member) for member in population], dtype=float)

    for _ in range(int(generations)):
        for idx in range(population_count):
            choices = [choice for choice in range(population_count) if choice != idx]
            a_idx, b_idx, c_idx = rng.choice(choices, size=3, replace=False)
            mutant = population[a_idx] + float(mutation) * (population[b_idx] - population[c_idx])
            mutant = _clip_to_bounds(mutant, lower_bounds, upper_bounds)
            crossover_mask = rng.random(dimensions) < float(crossover)
            if not crossover_mask.any():
                crossover_mask[int(rng.integers(0, dimensions))] = True
            trial = np.where(crossover_mask, mutant, population[idx])
            trial_score = float(objective_function(trial))
            if trial_score <= scores[idx]:
                population[idx] = trial
                scores[idx] = trial_score

    return population[int(np.argmin(scores))]


def _build_recommendation_row(method, variable, current_value, recommended_value, aggregate_improvement, current_scalar, recommended_scalar):
    movement = float(recommended_value) - float(current_value)
    return {
        "Method": method,
        "Variable": variable,
        "Current Value": float(current_value),
        "Recommended Value": float(recommended_value),
        "Change": movement,
        "Expected Improvement": float(aggregate_improvement),
        "Expected Improvement %": float(aggregate_improvement / max(abs(current_scalar), 1e-9) * 100.0),
        "Current Objective Score": float(current_scalar),
        "Recommended Objective Score": float(recommended_scalar),
    }


def generate_prescriptive_optimization_report(
    dataframe,
    control_variables=PRESCRIPTIVE_OPTIMIZATION_VARIABLES,
    objectives=PRESCRIPTIVE_OPTIMIZATION_OBJECTIVES,
    report_path=OPTIMIZATION_REPORT_PATH,
):
    """Generate Bayesian Optimization and Differential Evolution recommendations.

    The report varies only the requested controllable variables and constrains
    every recommendation to the exact observed historical min/max envelope.
    """
    optimization_df, missing_columns = _prepare_prescriptive_optimization_data(dataframe, control_variables, objectives)
    os.makedirs(os.path.dirname(report_path), exist_ok=True)

    if optimization_df is None or len(optimization_df) < 8:
        reason = (
            "Missing required columns: " + ", ".join(missing_columns)
            if missing_columns
            else "Not enough complete numeric historical rows for optimization."
        )
        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
            pd.DataFrame([{
                "status": "skipped",
                "reason": reason,
                "required_control_variables": ", ".join(control_variables),
                "required_objectives": ", ".join(objectives),
                "created_at_utc": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            }]).to_excel(writer, sheet_name="Summary", index=False)
        print(f"Prescriptive optimization skipped: {reason}")
        print(f"Saved optimization report: {report_path}")
        return report_path

    X = optimization_df[list(control_variables)].to_numpy(dtype=float)
    y_objectives = optimization_df[list(objectives)]
    objective_min, objective_range = _objective_normalization_parameters(y_objectives)
    scalar_y = _scalar_quality_objective(y_objectives, objective_min, objective_range)
    lower_bounds = optimization_df[list(control_variables)].min(axis=0).to_numpy(dtype=float)
    upper_bounds = optimization_df[list(control_variables)].max(axis=0).to_numpy(dtype=float)
    current_row = optimization_df.iloc[-1]
    current_controls = current_row[list(control_variables)].to_numpy(dtype=float)

    surrogate = ExtraTreesRegressor(n_estimators=100, random_state=model_seed, min_samples_leaf=2, n_jobs=1)
    surrogate.fit(X, y_objectives.to_numpy(dtype=float))

    current_scalar = float(_predict_scalar_quality(current_controls, surrogate, objective_min, objective_range)[0])
    objective_function = lambda candidate: float(
        _predict_scalar_quality(_clip_to_bounds(candidate, lower_bounds, upper_bounds), surrogate, objective_min, objective_range)[0]
    )

    bayesian_controls = _bayesian_optimize_controls(
        X,
        scalar_y,
        lower_bounds,
        upper_bounds,
        seed=model_seed,
    )
    de_controls = _differential_evolution_optimize_controls(
        objective_function,
        lower_bounds,
        upper_bounds,
        seed=model_seed,
    )

    recommendation_rows = []
    prediction_rows = []
    methods = OrderedDict([
        ("Bayesian Optimization", bayesian_controls),
        ("Differential Evolution", de_controls),
    ])
    for method, recommended_controls in methods.items():
        recommended_scalar = float(objective_function(recommended_controls))
        aggregate_improvement = current_scalar - recommended_scalar
        predicted_current_objectives = surrogate.predict(current_controls.reshape(1, -1)).reshape(-1)
        predicted_recommended_objectives = surrogate.predict(recommended_controls.reshape(1, -1)).reshape(-1)
        for variable, current_value, recommended_value in zip(control_variables, current_controls, recommended_controls):
            recommendation_rows.append(
                _build_recommendation_row(
                    method,
                    variable,
                    current_value,
                    recommended_value,
                    aggregate_improvement,
                    current_scalar,
                    recommended_scalar,
                )
            )
        for objective_name, current_prediction, recommended_prediction in zip(
            objectives,
            predicted_current_objectives,
            predicted_recommended_objectives,
        ):
            prediction_rows.append({
                "Method": method,
                "Objective": objective_name,
                "Current Predicted Value": float(current_prediction),
                "Recommended Predicted Value": float(recommended_prediction),
                "Expected Improvement": float(current_prediction - recommended_prediction),
                "Expected Improvement %": float((current_prediction - recommended_prediction) / max(abs(current_prediction), 1e-9) * 100.0),
            })

    limits_rows = [
        {
            "Variable": variable,
            "Observed Min": float(low),
            "Observed Max": float(high),
            "Current Value": float(current),
        }
        for variable, low, high, current in zip(control_variables, lower_bounds, upper_bounds, current_controls)
    ]
    summary_rows = [{
        "status": "completed",
        "created_at_utc": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
        "historical_rows_used": int(len(optimization_df)),
        "control_variables": ", ".join(control_variables),
        "objectives_minimized": ", ".join(objectives),
        "constraint": "Recommended values are clipped to, and selected inside, observed historical min/max limits.",
        "current_objective_score": current_scalar,
    }]

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        pd.DataFrame(recommendation_rows).to_excel(writer, sheet_name="Recommendations", index=False)
        pd.DataFrame(prediction_rows).to_excel(writer, sheet_name="Objective Forecasts", index=False)
        pd.DataFrame(limits_rows).to_excel(writer, sheet_name="Historical Limits", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="D9EAD3")
        for worksheet in workbook.worksheets:
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

    print(f"Saved optimization report: {report_path}")
    return report_path


def print_refinery_variable_groups():
    """Explain the refinery feature groups and the interactive override."""
    print_divider("=")
    print("Industrial refinery variable logic")
    print_divider("-")
    print("EARLY_VARIABLES (available early): " + ", ".join(EARLY_VARIABLES))
    print("CONTROL_VARIABLES (operator-adjustable): " + ", ".join(CONTROL_VARIABLES))
    print("TARGET_VARIABLES (future quality outputs): " + ", ".join(TARGET_VARIABLES))
    print("Default safe inputs are EARLY_VARIABLES + CONTROL_VARIABLES.")
    print("Interactive mode now lets you choose any non-output column from the active raw dataset CSV as an input.")
    print("Safety rule: a column selected as an output target is always blocked from X.")
    print_divider("=")


def _format_industrial_value(value):
    """Format values for compact refinery-control demo tables."""
    if isinstance(value, (int, float, np.integer, np.floating)):
        return f"{float(value):,.3f}"
    try:
        return f"{float(value):,.3f}"
    except (TypeError, ValueError):
        return str(value)


def _risk_badge(risk_level):
    """Return a professional text badge for LOW/MEDIUM/HIGH risk."""
    risk_level = str(risk_level).upper()
    badges = {
        "LOW": "LOW - Normal operating band",
        "MEDIUM": "MEDIUM - Watch and correct",
        "HIGH": "HIGH - Immediate operator attention",
    }
    return badges.get(risk_level, risk_level)



def _operator_expected_improvement(recommendation):
    """Return aggregate expected quality gain (positive means lower predicted objective)."""
    current_prediction = recommendation.get("current_prediction", {}) or {}
    future_quality = recommendation.get("predicted_future_quality", {}) or {}
    target_weights = recommendation.get("target_weights", {}) or {}
    improvements = []
    for target, weight in target_weights.items():
        if target not in current_prediction or target not in future_quality:
            continue
        improvements.append((float(current_prediction[target]) - float(future_quality[target])) * float(weight))
    if improvements:
        return float(sum(improvements) / max(sum(abs(float(w)) for w in target_weights.values()), 1e-9))

    for target, before in current_prediction.items():
        if target in future_quality:
            improvements.append(float(before) - float(future_quality[target]))
    return float(sum(improvements) / len(improvements)) if improvements else 0.0


def _operator_confidence(recommendation):
    """Estimate recommendation confidence from model risk and search coverage."""
    risk_level = str(recommendation.get("risk_prediction", {}).get("risk_level", "MEDIUM")).upper()
    base = {"LOW": 0.86, "MEDIUM": 0.70, "HIGH": 0.52}.get(risk_level, 0.60)
    searched = int(recommendation.get("searched_candidates", 0) or 0)
    if searched >= 1000:
        base += 0.04
    elif searched < 25:
        base -= 0.08
    return max(0.05, min(base, 0.95))


def _autofit_excel_columns(writer, sheet_name):
    """Resize worksheet columns after a pandas export."""
    worksheet = writer.sheets.get(sheet_name)
    if worksheet is None:
        return
    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max([len(value) for value in values] + [10]) + 2, 60)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")


def save_operator_report(recommendation, report_path=OPERATOR_REPORT_PATH):
    """Write the operator-facing refinery recommendation workbook.

    The workbook gives operators the requested side-by-side view of the current
    set-points, AI-recommended set-points, predicted quality change, risk level,
    and confidence, plus supporting quality and advisory sheets.
    """
    os.makedirs(os.path.dirname(report_path), exist_ok=True)

    current_settings = recommendation.get("current_settings", {}) or {}
    recommended_settings = recommendation.get("recommended_settings", {}) or {}
    current_prediction = recommendation.get("current_prediction", {}) or {}
    future_quality = recommendation.get("predicted_future_quality", {}) or {}
    risk = recommendation.get("risk_prediction", {}) or {}
    current_risk = recommendation.get("current_risk_prediction", {}) or {}
    control_ranges = recommendation.get("control_ranges", {}) or {}
    improvement = _operator_expected_improvement(recommendation)
    confidence = _operator_confidence(recommendation)
    risk_level = risk.get("risk_level", "UNKNOWN")

    setting_rows = []
    for variable in OPERATOR_RECOMMENDATION_VARIABLES:
        if variable not in recommended_settings and variable not in current_settings:
            continue
        current_value = current_settings.get(variable)
        recommended_value = recommended_settings.get(variable)
        change = None
        if current_value is not None and recommended_value is not None:
            change = float(recommended_value) - float(current_value)
        low, high = control_ranges.get(variable, (None, None))
        setting_rows.append({
            "Setting": variable,
            "Current": current_value,
            "Recommended": recommended_value,
            "Change": change,
            "Expected Quality Improvement": improvement,
            "Risk Level": risk_level,
            "Confidence": confidence,
            "Search Low": low,
            "Search High": high,
        })

    quality_rows = []
    for target, current_value in current_prediction.items():
        recommended_value = future_quality.get(target)
        quality_rows.append({
            "Quality Target": target,
            "Current Predicted Quality": current_value,
            "Recommended Predicted Quality": recommended_value,
            "Expected Quality Improvement": (
                float(current_value) - float(recommended_value)
                if recommended_value is not None else None
            ),
            "Current Risk Level": current_risk.get("risk_level", "UNKNOWN"),
            "Recommended Risk Level": risk_level,
            "Confidence": confidence,
        })

    summary_rows = [{
        "Created At UTC": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
        "Current Predicted Quality": json.dumps(current_prediction, ensure_ascii=False, default=float),
        "Recommended Predicted Quality": json.dumps(future_quality, ensure_ascii=False, default=float),
        "Expected Quality Improvement": improvement,
        "Risk Level": risk_level,
        "Confidence": confidence,
        "Candidate Simulations": recommendation.get("searched_candidates"),
        "Report Path": report_path,
    }]

    risk_driver_rows = risk.get("risk_drivers", []) or []
    warning_rows = [{"Operator Advisory": warning} for warning in risk.get("operator_warnings", [])]

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(setting_rows).to_excel(writer, sheet_name="Recommended Settings", index=False)
        pd.DataFrame(quality_rows).to_excel(writer, sheet_name="Predicted Quality", index=False)
        pd.DataFrame(risk_driver_rows).to_excel(writer, sheet_name="Risk Drivers", index=False)
        pd.DataFrame(warning_rows).to_excel(writer, sheet_name="Operator Advisory", index=False)
        for sheet_name in writer.sheets:
            _autofit_excel_columns(writer, sheet_name)

    print(f"Saved operator report: {report_path}")
    return report_path


def print_industrial_operator_demo(recommendation, input_features, output_features):
    """Print a realistic step-by-step operator demonstration.

    The section mirrors the requested plant workflow: current conditions are
    entered, the AI predicts future sugar quality, the AI recommends controllable
    variables, then it prints expected quality and risk in a readable shift-log
    format.
    """
    current_inputs = recommendation.get("full_candidate_inputs", {}).copy()
    current_settings = recommendation.get("current_settings", {})
    for variable, value in current_settings.items():
        current_inputs[variable] = value

    recommended = recommendation.get("recommended_settings", {})
    current_prediction = recommendation.get("current_prediction", {})
    future_quality = recommendation.get("predicted_future_quality", {})
    current_risk = recommendation.get("current_risk_prediction", {})
    future_risk = recommendation.get("risk_prediction", {})
    control_ranges = recommendation.get("control_ranges", {})

    print("\n")
    print_divider("=", 76)
    print("INDUSTRIAL OPERATOR DEMO - AI REFINERY QUALITY ADVISOR")
    print_divider("=", 76)
    print("Scenario: shift operator enters live refinery conditions and requests an")
    print("AI advisory for future sugar quality, optimal controllable set-points,")
    print("expected future quality, and operating risk.")

    print("\n1) OPERATOR ENTERS CURRENT REFINERY CONDITIONS")
    print_divider("-", 76)
    condition_rows = []
    for feature in input_features:
        if feature not in current_inputs:
            continue
        group = "Controllable" if feature in CONTROL_VARIABLES else "Early/process"
        condition_rows.append([group, feature, _format_industrial_value(current_inputs[feature])])
    print(tabulate(condition_rows, headers=["Group", "Variable", "Current value"], tablefmt="github"))

    print("\n2) AI PREDICTS FUTURE SUGAR QUALITY AT CURRENT CONDITIONS")
    print_divider("-", 76)
    current_quality_rows = []
    for target in output_features:
        if target in current_prediction:
            current_quality_rows.append([target, _format_industrial_value(current_prediction[target])])
    print(tabulate(current_quality_rows, headers=["Future quality target", "Predicted value"], tablefmt="github"))
    print(f"Current-condition risk: {_risk_badge(current_risk.get('risk_level', 'UNKNOWN'))}")

    print("\n3) AI RECOMMENDS OPTIMAL CONTROLLABLE VARIABLES")
    print_divider("-", 76)
    recommendation_rows = []
    for variable, recommended_value in recommended.items():
        low, high = control_ranges.get(variable, (None, None))
        safe_range = "N/A" if low is None or high is None else f"{low:.3f} - {high:.3f}"
        current_value = current_settings.get(variable, current_inputs.get(variable))
        delta = float(recommended_value) - float(current_value)
        recommendation_rows.append([
            variable,
            _format_industrial_value(current_value),
            _format_industrial_value(recommended_value),
            f"{delta:+.3f}",
            safe_range,
        ])
    print(tabulate(
        recommendation_rows,
        headers=["Control variable", "Current", "Recommended", "Change", "Search range"],
        tablefmt="github",
    ))

    print("\n4) AI PRINTS EXPECTED FUTURE QUALITY AFTER RECOMMENDATION")
    print_divider("-", 76)
    future_quality_rows = []
    for target in output_features:
        if target not in future_quality:
            continue
        before = current_prediction.get(target)
        after = future_quality[target]
        improvement = "N/A" if before is None else f"{float(before) - float(after):+.3f}"
        future_quality_rows.append([
            target,
            _format_industrial_value(before) if before is not None else "N/A",
            _format_industrial_value(after),
            improvement,
        ])
    print(tabulate(
        future_quality_rows,
        headers=["Future quality target", "Current prediction", "Expected with AI set-points", "Improvement"],
        tablefmt="github",
    ))

    print("\n5) AI PRINTS INDUSTRIAL RISK LEVEL")
    print_divider("-", 76)
    print(f"Recommended-condition risk: {_risk_badge(future_risk.get('risk_level', 'UNKNOWN'))}")
    risk_driver_rows = []
    for item in future_risk.get("risk_drivers", []):
        risk_driver_rows.append([
            item.get("target"),
            _format_industrial_value(item.get("predicted_value")),
            item.get("risk_level"),
            item.get("reason"),
        ])
    if risk_driver_rows:
        print(tabulate(
            risk_driver_rows,
            headers=["Risk driver", "Predicted value", "Level", "Reason"],
            tablefmt="github",
        ))
    print("Operator advisory:")
    for warning in future_risk.get("operator_warnings", []):
        print(f" - {warning}")
    print(f"AI candidate simulations reviewed: {recommendation.get('searched_candidates', 'N/A')}")
    print_divider("=", 76)


def list_saved_runs(saved_dir=SAVED_RUNS_DIR):
    os.makedirs(saved_dir, exist_ok=True)
    saved = []
    for file_name in sorted(os.listdir(saved_dir)):
        if not file_name.endswith(".json"):
            continue
        path = os.path.join(saved_dir, file_name)
        try:
            with open(path, "r", encoding="utf-8") as fp:
                payload = json.load(fp)
            payload["_path"] = path
            saved.append(payload)
        except (OSError, json.JSONDecodeError):
            continue
    return saved


def prompt_saved_run_choice():
    while True:
        saved = list_saved_runs()
        if not saved:
            return None, False
        print("\nSaved runs:")
        table_rows = []
        for idx, run in enumerate(saved, start=1):
            table_rows.append(
                [
                    idx,
                    run.get("display_name", "unknown"),
                    run.get("best_r2", "N/A"),
                    run.get("saved_at", "N/A"),
                    "yes" if run.get("has_weights") else "no",
                ]
            )
        print(tabulate(table_rows, headers=["#", "Saved model", "R2", "Date", "Weights"], tablefmt="github"))
        print("Tip: type delete_<index> (example: delete_3) to remove a saved model/checkpoint.")
        choice = input("Select a saved run by index or press Enter for new run [new]: ").strip()
        if not choice:
            return None, False
        if choice.lower().startswith("delete_"):
            index_str = choice.split("_", 1)[1].strip()
            if not index_str.isdigit():
                print("Invalid delete format. Use delete_<index> such as delete_3.")
                continue
            idx = int(index_str)
            if idx < 1 or idx > len(saved):
                print("Invalid delete index.")
                continue
            selected_for_delete = saved[idx - 1]
            summary_path = selected_for_delete.get("_path")
            weights_path = selected_for_delete.get("weights_path")
            try:
                if summary_path and os.path.isfile(summary_path):
                    os.remove(summary_path)
                    print(f"Deleted summary file: {summary_path}")
                if weights_path and os.path.isfile(weights_path):
                    os.remove(weights_path)
                    print(f"Deleted checkpoint file: {weights_path}")
            except OSError as err:
                print(f"Could not delete saved model files ({err}).")
            continue
        if not choice.isdigit() or int(choice) < 1 or int(choice) > len(saved):
            print("Invalid choice. Starting a new run.")
            return None, False
        selected = saved[int(choice) - 1]
        has_weights = bool(selected.get("has_weights"))
        if has_weights:
            mode = input(
                "Selected saved run has model weights. Continue by optimizing [weights|inputs|both] [weights]: "
            ).strip().lower() or "weights"
        else:
            mode = input(
                "Selected saved run has configs only (no weights). Continue by optimizing [weights|inputs|both] [weights]: "
            ).strip().lower() or "weights"
        return selected, mode


def save_run_summary(
    model_name,
    inputs,
    outputs,
    best_r2,
    epoch_candidates=None,
    hidden_size_groups=None,
    repeat_count=None,
    optimization_scope=None,
    weights_path=None,
    saved_dir=SAVED_RUNS_DIR,
):
    os.makedirs(saved_dir, exist_ok=True)
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    display_name = f"{model_name}_in{len(inputs)}_out{len(outputs)}_R2-{best_r2:.2f}"
    payload = {
        "display_name": display_name,
        "model_name": model_name,
        "input_count": len(inputs),
        "output_count": len(outputs),
        "inputs": inputs,
        "outputs": outputs,
        "best_r2": round(float(best_r2), 2),
        "saved_at": now,
        "epoch_candidates": epoch_candidates if epoch_candidates is not None else [],
        "hidden_size_groups": hidden_size_groups if hidden_size_groups is not None else [],
        "repeat_count": int(repeat_count) if repeat_count is not None else None,
        "optimization_scope": optimization_scope,
        "weights_path": weights_path,
        "has_weights": bool(weights_path and os.path.isfile(weights_path)),
    }
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", display_name)
    summary_path = os.path.join(saved_dir, f"{safe_name}.json")
    with open(summary_path, "w", encoding="utf-8") as fp:
        json.dump(payload, fp, indent=2, ensure_ascii=False)
    print(f"Saved run summary: {summary_path}")


def normalize_hidden_size_groups(raw_groups, fallback_groups):
    """Return valid hidden-size groups as list[list[int]]."""
    parsed = []
    if isinstance(raw_groups, list):
        for group in raw_groups:
            if isinstance(group, list):
                cleaned = [int(v) for v in group if str(v).isdigit() and int(v) > 0]
                if cleaned:
                    parsed.append(cleaned)
    return parsed if parsed else [list(group) for group in fallback_groups]

list_epochs = [50, 100, 200, 300]
best_epochs = 100

exp_df = pd.DataFrame()

num_repeats = 5
# num_repeats shows the number of times to repeat the experiment to get its average values
# changing model_seed generates different results for each run for each model
# The same model_seed produces the same results in each run which is good for reporducability of experiments
# Dont change it if you want to reproduce the same results
data_seed = 123 # it is used for random_state of splitting data into source and train sets
# changing it creates different source and train sets.
# Since the number of data is low changing it can largely affect the results
model_seed = 123 # it is used for random_state of models
def set_model_seed(model_seed):
    torch.manual_seed(model_seed)
    np.random.seed(model_seed)

learning_rate = 0.0005
# Industrially conservative ANN weight learning rate. 0.05 was too aggressive
# for a small, standardized tabular dataset and can overshoot narrow minima.
input_learning_rate = 0.00005
# Input optimization is more sensitive than weight optimization because it moves
# the normalized training data directly, so keep it an order of magnitude lower.
ann_weight_decay = 1e-4
# L2 regularization reduces small-sample overfitting without increasing model size.
early_stopping_patience = 20
early_stopping_min_delta = 1e-4
validation_fraction = 0.2
lr_plateau_patience = 8
lr_plateau_factor = 0.5
min_learning_rate = 1e-5
max_gradient_norm = 0.5
max_input_delta = 1.5
# Conservative training guards for stable convergence on small industrial data.
#hidden_size1 = 10
hidden_size1 = 32
hidden_size2 = 16

# the number of neurons in hidden layers

# https://alexlenail.me/NN-SVG/
# use the site above to draw the following network
#
hidden_sizes = [32, 16]
# nn.ReLU(), nn.Tanh(), nn.Identity()

list_hidden_sizes = [[16], [32], [64], [16, 8], [32, 16], [48, 24], [64, 32], [64, 32, 16]]
normalization_type = "standard_scaler"

EXTRA_TREES_RANDOM_SEARCH_ITERATIONS = 40
EXTRA_TREES_RANDOM_SEARCH_CV_FOLDS = 5
MODEL_EVALUATION_CV_FOLDS = 5


def make_extra_trees_random_search(seed):
    """Build a tuned ExtraTrees regressor while preserving the model menu name."""
    base_model = ExtraTreesRegressor(random_state=seed, n_jobs=-1)
    param_distributions = {
        "n_estimators": [300, 500, 800, 1000],
        "max_depth": [None, 4, 6, 8, 10, 12, 16],
        "min_samples_split": [2, 3, 4, 5, 8, 10],
        "min_samples_leaf": [1, 2, 3, 4, 6],
        "max_features": ["sqrt", "log2", 0.5, 0.7, 0.9, 1.0],
        "bootstrap": [False, True],
    }
    return RandomizedSearchCV(
        estimator=base_model,
        param_distributions=param_distributions,
        n_iter=EXTRA_TREES_RANDOM_SEARCH_ITERATIONS,
        scoring=make_scorer(r2_score, multioutput="uniform_average"),
        cv=EXTRA_TREES_RANDOM_SEARCH_CV_FOLDS,
        random_state=seed,
        n_jobs=-1,
        refit=True,
    )


def make_catboost_regressor(seed):
    """Build a quiet CatBoost model for small tabular refinery datasets."""
    return CatBoostRegressor(
        iterations=1200,
        learning_rate=0.03,
        depth=6,
        l2_leaf_reg=5.0,
        loss_function="RMSE",
        random_seed=seed,
        verbose=False,
        allow_writing_files=False,
    )


SKLEARN_MODEL_FACTORIES = {
    "RandomForestRegressor": lambda seed: RandomForestRegressor(
        n_estimators=300, random_state=seed
    ),
    "ExtraTreesRegressor": make_extra_trees_random_search,
    "GradientBoostingRegressor": lambda seed: GradientBoostingRegressor(random_state=seed),
    "GBMRegressor": lambda seed: GradientBoostingRegressor(random_state=seed),
    "HistGradientBoostingRegressor": lambda seed: HistGradientBoostingRegressor(
        random_state=seed
    ),
    "AdaBoostRegressor": lambda seed: AdaBoostRegressor(random_state=seed),
    "DecisionTreeRegressor": lambda seed: DecisionTreeRegressor(random_state=seed),
    "KNeighborsRegressor": lambda seed: KNeighborsRegressor(n_neighbors=5),
    "SVR_RBF": lambda seed: SVR(kernel="rbf"),
    "LinearRegression": lambda seed: LinearRegression(),
    "Ridge": lambda seed: Ridge(random_state=seed),
    "Lasso": lambda seed: Lasso(random_state=seed),
    "ElasticNet": lambda seed: ElasticNet(random_state=seed),
}

if XGBRegressor is not None:
    SKLEARN_MODEL_FACTORIES["XGBoostRegressor"] = lambda seed: XGBRegressor(
        n_estimators=300,
        random_state=seed,
        objective="reg:squarederror",
    )
else:
    print("Optional dependency not found: xgboost. Skipping XGBoostRegressor.")

if LGBMRegressor is not None:
    SKLEARN_MODEL_FACTORIES["LightGBMRegressor"] = lambda seed: LGBMRegressor(
        n_estimators=300,
        random_state=seed,
        verbose=-1,
    )
else:
    print("Optional dependency not found: lightgbm. Skipping LightGBMRegressor.")

if CatBoostRegressor is not None:
    SKLEARN_MODEL_FACTORIES["CatBoostRegressor"] = make_catboost_regressor
else:
    print("Optional dependency not found: catboost. Skipping CatBoostRegressor.")


def _expand_index_token(token, options_len, one_based=False):
    """Expand a token like '3' or '1-4' into zero-based indexes."""
    if "-" in token:
        bounds = token.split("-", 1)
        if len(bounds) != 2 or not bounds[0].isdigit() or not bounds[1].isdigit():
            raise ValueError(f"Invalid range '{token}'. Use syntax like 1-4.")
        start = int(bounds[0])
        end = int(bounds[1])
        if start > end:
            raise ValueError(f"Invalid range '{token}'. Start must be <= end.")
        raw_indexes = list(range(start, end + 1))
    else:
        if not token.isdigit():
            raise ValueError(f"Invalid input '{token}'. Please use indexes or ranges like 1-4.")
        raw_indexes = [int(token)]

    indexes = [index - 1 for index in raw_indexes] if one_based else raw_indexes
    min_valid = 1 if one_based else 0
    max_valid = options_len if one_based else options_len - 1
    for raw_index, index in zip(raw_indexes, indexes):
        if index < 0 or index >= options_len:
            raise ValueError(
                f"Invalid selection '{raw_index}'. Valid range is {min_valid} to {max_valid}."
            )
    return indexes


def parse_multi_select(answer, options, allow_all=True, one_based=False):
    """Parse indexes/ranges/names and return selected option values.

    Supported syntax examples:
    - 0 2 4
    - 1-5
    - 1,3,6-9
    - !0 !4-6 (exclude indexes/ranges from current pool)
    - raw_syrup_brix,standard_liquor_color (feature names)
    """
    if not answer:
        return None if allow_all else []

    answer = answer.strip().lower()
    if allow_all and answer == "all":
        return None

    include_indexes = set()
    exclude_indexes = set()
    tokens = [token for token in answer.replace(",", " ").split() if token]
    if not tokens:
        return None if allow_all else []

    for token in tokens:
        is_exclude = token.startswith("!")
        clean = token[1:] if is_exclude else token
        if clean == "all":
            if not allow_all:
                raise ValueError("'all' is not allowed for this selection.")
            include_indexes.update(range(len(options)))
            continue

        if clean.isdigit() or ("-" in clean and clean.replace("-", "").isdigit()):
            indexes = _expand_index_token(clean, len(options), one_based=one_based)
        else:
            matching_indexes = [
                idx for idx, option in enumerate(options)
                if str(option).lower() == clean
            ]
            if not matching_indexes:
                raise ValueError(
                    f"Invalid input '{token}'. Please use indexes, ranges like 1-4, "
                    "or exact feature names from the menu."
                )
            indexes = matching_indexes
        if is_exclude:
            exclude_indexes.update(indexes)
        else:
            include_indexes.update(indexes)

    if include_indexes:
        final_indexes = sorted(include_indexes.difference(exclude_indexes))
    else:
        # Exclude-only syntax means "all except excluded".
        if exclude_indexes:
            final_indexes = [i for i in range(len(options)) if i not in exclude_indexes]
        elif allow_all:
            return None
        else:
            final_indexes = []

    if not final_indexes:
        raise ValueError("Selection is empty after applying include/exclude filters.")

    selected_values = [options[index] for index in final_indexes]
    return selected_values



def selectable_output_features_from_csv(data):
    """Return output choices from every active raw dataset column, targets first."""
    columns = list(data.columns)
    target_first = [column for column in columns if column in TARGET_VARIABLES]
    remaining = [column for column in columns if column not in target_first]
    return target_first + remaining


def selectable_input_features_from_csv(data, selected_output_features):
    """Return every CSV column that is not currently selected as an output."""
    selected_outputs = set(selected_output_features or [])
    return [column for column in data.columns if column not in selected_outputs]


def optional_input_overrides_for_selection(selected_input_features, selected_output_features):
    """Whitelist explicitly selected non-output inputs for this interactive run.

    read_data.py still prevents selected outputs from being used as inputs.  This
    opt-in list lets run.py honor the user's raw dataset input choices instead
    of silently reducing the selection to refinery default groups.
    """
    selected_outputs = set(selected_output_features or [])
    return [feature for feature in selected_input_features if feature not in selected_outputs]

def print_future_quality_variables_menu():
    print("\n" + "-" * 50)
    print("Future Quality Variables")
    for idx, variable_name in enumerate(FUTURE_QUALITY_INPUT_CANDIDATES, start=1):
        print(f"{idx}) {variable_name}")
    print("-" * 50)


def append_future_quality_input_candidates(
    active_input_features,
    available_columns,
    selected_output_features=None,
    prompt_label="Select future-quality variables to add as candidate input features",
):
    """Prompt for optional future-quality inputs and append valid, unique selections.

    Future-quality variables are normally targets. This opt-in prompt lets the
    user include available non-target quality measurements as normal model
    inputs for future-quality prediction while preserving feature order and
    preventing duplicates.
    """
    active_input_features = list(active_input_features)
    available_columns = list(available_columns)
    selected_output_features = set(selected_output_features or [])

    print_future_quality_variables_menu()
    answer = input(
        f"{prompt_label} (indexes/ranges), [all], or Enter for none [none]: "
    ).strip()
    if not answer:
        print("No additional future-quality input variables selected.")
        return active_input_features, []

    try:
        selected_variables = parse_multi_select(
            answer,
            FUTURE_QUALITY_INPUT_CANDIDATES,
            allow_all=True,
            one_based=True,
        )
    except ValueError as err:
        print(f"Invalid future-quality input selection: {err}")
        exit()

    if selected_variables is None:
        selected_variables = list(FUTURE_QUALITY_INPUT_CANDIDATES)

    appended_variables = []
    skipped_missing = []
    skipped_outputs = []
    skipped_duplicates = []

    for variable_name in selected_variables:
        if variable_name in selected_output_features:
            skipped_outputs.append(variable_name)
            continue
        if variable_name not in available_columns:
            skipped_missing.append(variable_name)
            continue
        if variable_name in active_input_features:
            skipped_duplicates.append(variable_name)
            continue
        active_input_features.append(variable_name)
        appended_variables.append(variable_name)

    if appended_variables:
        print("Added future-quality candidate input feature(s): " + ", ".join(appended_variables))
    else:
        print("No future-quality variables were added to the input feature list.")

    if skipped_outputs:
        print(
            "Warning: selected future-quality variable(s) are current output target(s) "
            "and were skipped to avoid target leakage: " + ", ".join(skipped_outputs)
        )
    if skipped_missing:
        print(
            "Warning: selected future-quality variable(s) are not present in X_train.columns "
            "and were skipped: " + ", ".join(skipped_missing)
        )
    if skipped_duplicates:
        print(
            "Already selected as input feature(s); duplicates were ignored: "
            + ", ".join(skipped_duplicates)
        )

    return active_input_features, appended_variables


def print_selection_guide():
    print("\nSelection guide:")
    print("- Single indexes: 1 3 5")
    print("- Range syntax: 1-9")
    print("- Mixed syntax: 1 2-4 7")
    print("- Commas are also allowed: 1,2-4,7")
    print("- Exclude indexes/ranges with ! : !1 !5-8")
    print("- Exact feature names are accepted too: raw_syrup_brix,standard_liquor_color")
    print("- In auto mode, exclude-only means all features except those indexes.")


def prompt_temporal_options(input_features):
    print("\nTemporal/sequential feature options:")
    print("Enter indexes/ranges for sequential columns if any (empty for none).")
    print("Optional grouped dependencies: [2 4] [3 11 13] or 2>4>6,3>11>13")
    print("\n".join([str(i) + ")" + name for i, name in enumerate(input_features, start=1)]))
    seq_answer = input("Sequential feature columns/groups [none]: ").strip()
    if not seq_answer:
        return [], [], False, 1, False, False, False, False, 3, False, 3

    try:
        sequential_features, sequential_groups = parse_sequential_layout(seq_answer, input_features, one_based=True)
    except ValueError as err:
        print(f"Invalid sequential feature selection: {err}")
        exit()

    add_differences = input("Add difference features for sequential columns? [n]: ").strip().lower() in ("y", "yes")
    difference_order = 1
    if add_differences:
        difference_order = int(input("Difference order [1]: ").strip() or "1")

    add_ratio_features = input("Add ratio features (x_t / x_{t-1})? [n]: ").strip().lower() in ("y", "yes")
    add_acceleration_features = input(
        "Add acceleration features (x_t - 2*x_{t-1} + x_{t-2})? [n]: "
    ).strip().lower() in ("y", "yes")
    add_normalized_change_features = input(
        "Add normalized change features ((x_t - x_{t-1}) / x_{t-1})? [n]: "
    ).strip().lower() in ("y", "yes")

    create_rnn_windows = input("Add lag-window features for sequential columns (lag_1, lag_2, lag_3 by default)? [n]: ").strip().lower() in ("y", "yes")
    rnn_window_size = 3
    if create_rnn_windows:
        rnn_window_size = int(input("Largest lag step [3]: ").strip() or "3")

    add_rolling_dynamics = input(
        "Add rolling process dynamics (mean/std/range/slope)? [n]: "
    ).strip().lower() in ("y", "yes")
    rolling_window = 3
    if add_rolling_dynamics:
        rolling_window = int(input("Rolling dynamics window [3]: ").strip() or "3")

    return (
        sequential_features,
        sequential_groups,
        add_differences,
        difference_order,
        add_ratio_features,
        add_acceleration_features,
        add_normalized_change_features,
        create_rnn_windows,
        rnn_window_size,
        add_rolling_dynamics,
        rolling_window,
    )


def parse_sequential_layout(answer, options, one_based=False):
    """Parse independent sequential columns and optional ordered dependency groups."""
    grouped_layout = (
        "[" in answer or "]" in answer or "(" in answer or ")" in answer or ">" in answer
    )
    if not grouped_layout:
        sequential_features = parse_multi_select(answer, options, allow_all=False, one_based=one_based)
        return sequential_features, []

    groups = []
    bracket_groups = re.findall(r"[\[\(]([^\]\)]+)[\]\)]", answer)
    raw_groups = bracket_groups if bracket_groups else [item.strip() for item in answer.split(",") if item.strip()]

    for raw_group in raw_groups:
        normalized = raw_group.replace(">", " ")
        group_values = parse_multi_select(normalized, options, allow_all=False, one_based=one_based)
        if len(group_values) < 2:
            raise ValueError("Each sequential group must contain at least two columns.")
        groups.append(group_values)

    flat_features = []
    for group in groups:
        for feature in group:
            if feature not in flat_features:
                flat_features.append(feature)
    return flat_features, groups


def prepare_or_reuse_data(dataset_path="convert/stclean.csv", prep_folder="prep_data"):
    print("====================================")
    print("DATASET PATH: " + dataset_path)
    print("====================================")
    print(f"[path-debug] run.py file: {os.path.abspath(__file__)}")
    dataset_path = print_path_debug("raw dataset CSV", dataset_path)
    prep_folder = print_path_debug("prep_data folder", prep_folder)
    print(f"[path-debug] Effective dataset path: {dataset_path}")
    print(f"[path-debug] Effective prep_data folder: {prep_folder}")

    use_auto_feature_selection = False
    reused_prep_data = False
    print_refinery_variable_groups()

    if prep_data_exists(prep_folder):
        prep_x_train_path = os.path.join(prep_folder, "X_train.csv")
        prep_y_train_path = os.path.join(prep_folder, "y_train.csv")
        existing_all_inputs = pd.read_csv(prep_x_train_path, nrows=0).columns.tolist()
        existing_outputs = pd.read_csv(prep_y_train_path, nrows=0).columns.tolist()
        leaked_existing_inputs = find_leakage_columns(
            existing_all_inputs,
            output_features=existing_outputs,
            optional_future_quality_inputs=FUTURE_QUALITY_INPUT_CANDIDATES,
        )
        if leaked_existing_inputs:
            print(
                "Existing prep_data contains future/disallowed inputs and will be rebuilt "
                f"to prevent target leakage: {leaked_existing_inputs}"
            )
        else:
            X_train_existing, _, y_train_existing, _ = read_prep_data(inputs=None, prep_folder=prep_folder)
            existing_inputs = X_train_existing.columns.tolist()
            print_numbered_feature_list("Prepared Input Features (prep_data)", existing_inputs, ANSI_GREEN)
            existing_inputs, optional_future_quality_inputs = append_future_quality_input_candidates(
                existing_inputs,
                existing_all_inputs,
                selected_output_features=existing_outputs,
            )
            if optional_future_quality_inputs:
                X_train_existing, _, y_train_existing, _ = read_prep_data(
                    inputs=existing_inputs,
                    prep_folder=prep_folder,
                    optional_future_quality_inputs=optional_future_quality_inputs,
                )
            print_numbered_feature_list("Prepared Output Features (prep_data)", existing_outputs, ANSI_BLUE)

            metadata = read_prep_metadata(prep_folder)
            if metadata:
                print("prep_data metadata:")
                print(metadata)

            print(
                "[path-debug] Complete prep_data files were found. If you continue with prep_data, "
                "the raw dataset CSV will not be read in this run."
            )
            reuse_answer = input(
                "Continue with these prepared refinery-safe inputs/outputs? [y]: "
            ).strip().lower()
            if reuse_answer in ("", "y", "yes"):
                print("[path-debug] prep_data reuse selected; skipping raw dataset CSV read.")
                reused_prep_data = True
                return X_train_existing, existing_outputs, use_auto_feature_selection, reused_prep_data
            print("[path-debug] prep_data reuse declined; raw dataset CSV flow will be used.")

            prep_train_path = os.path.join(prep_folder, "train.csv")
            prep_test_path = os.path.join(prep_folder, "test.csv")

            if os.path.isfile(prep_train_path) and os.path.isfile(prep_test_path):
                print("\nYou chose not to continue with current prep_data selection.")
                print("1) Reselect input features from prep_data/train.csv and prep_data/test.csv")
                print("2) Rebuild prep_data from the raw dataset and select input/output/sequential options again")
                reselect_mode = input("Choose option [1]: ").strip()
                if reselect_mode in ("", "1"):
                    train_df = pd.read_csv(prep_train_path)
                    test_df = pd.read_csv(prep_test_path)
                    available_inputs = [
                        column for column in train_df.columns if column not in set(existing_outputs)
                    ]
                    if not available_inputs:
                        print("No candidate input columns found. Falling back to full prepare flow.")
                    else:
                        print_numbered_feature_list(
                            "Input Features from prep_data/train.csv (prepared outputs excluded)",
                            available_inputs,
                            ANSI_GREEN,
                        )
                        reselect_answer = input(
                            "\nSelect one or several input features (indexes/ranges), or [all]: "
                        ).strip()
                        try:
                            selected_input_features = parse_multi_select(
                                reselect_answer,
                                available_inputs,
                                allow_all=True,
                                one_based=True,
                            )
                        except ValueError as err:
                            print(f"Invalid input feature selection: {err}")
                            exit()

                        resolved_inputs = (
                            available_inputs if selected_input_features is None else selected_input_features
                        )
                        resolved_inputs, _ = remove_leakage_inputs_for_training(
                            resolved_inputs,
                            output_features=existing_outputs,
                            context="prep_data_reselect",
                        )
                        if not resolved_inputs:
                            print("No input columns remain after automatic target-leakage removal.")
                            exit()
                        optional_future_quality_inputs = optional_input_overrides_for_selection(
                            resolved_inputs, existing_outputs
                        )
                        validate_model_inputs(
                            resolved_inputs,
                            output_features=existing_outputs,
                            optional_future_quality_inputs=optional_future_quality_inputs,
                        )
                        missing_inputs_in_test = [
                            col for col in resolved_inputs if col not in test_df.columns
                        ]
                        if missing_inputs_in_test:
                            print(
                                "Selected inputs are missing from prep_data/test.csv: "
                                + str(missing_inputs_in_test)
                            )
                            print("Please rebuild prep_data from the raw dataset.")
                            exit()

                        missing_outputs_in_test = [
                            col for col in existing_outputs if col not in test_df.columns
                        ]
                        if missing_outputs_in_test:
                            print(
                                "Prepared output columns are missing from prep_data/test.csv: "
                                + str(missing_outputs_in_test)
                            )
                            print("Please rebuild prep_data from the raw dataset.")
                            exit()

                        X_train_selected = train_df[resolved_inputs]
                        X_test_selected = test_df[resolved_inputs]
                        y_train_selected = train_df[existing_outputs]
                        y_test_selected = test_df[existing_outputs]
                        metadata = read_prep_metadata(prep_folder)
                        if metadata:
                            metadata["input_features"] = list(resolved_inputs)
                            metadata["refinery_variable_groups"] = refinery_variable_group_metadata()
                        save_data(
                            prep_folder,
                            X_train_selected,
                            X_test_selected,
                            y_train_selected,
                            y_test_selected,
                            metadata=metadata,
                        )
                        print("prep_data updated with the newly selected input features.")
                        reused_prep_data = True
                        return (
                            X_train_selected,
                            existing_outputs,
                            use_auto_feature_selection,
                            reused_prep_data,
                        )
                elif reselect_mode == "2":
                    print("Preparing data again from the raw dataset...")
                else:
                    print("Invalid choice. Preparing data again from the raw dataset...")
            else:
                print("prep_data/train.csv or prep_data/test.csv not found. Preparing data again from the raw dataset...")

    dataset_path = ensure_dataset_csv_exists(dataset_path)
    print(f"[path-debug] prep_data did not prevent raw CSV flow; reading dataset now: {dataset_path}")
    data = pd.read_csv(dataset_path)
    print_selection_guide()

    output_options = selectable_output_features_from_csv(data)
    print("\nOutput feature candidates from the active raw dataset CSV (TARGET_VARIABLES shown first):")
    print("\n".join([str(i) + ")" + name for i, name in enumerate(output_options, start=1)]))
    answer = input(
        "Select one or several output features (indexes/ranges/names) [first target candidate]:"
    )
    try:
        selected_output_features = (
            [output_options[0]]
            if not answer
            else parse_multi_select(answer, output_options, allow_all=False, one_based=True)
        )
    except ValueError as err:
        print(f"Invalid output feature selection: {err}")
        exit()

    available_input_features = selectable_input_features_from_csv(data, selected_output_features)
    if not available_input_features:
        print("No input columns are available after excluding selected output feature(s).")
        exit()
    print("\nAvailable model inputs from the active raw dataset CSV (selected outputs excluded):")
    print("\n".join([str(i) + ")" + name for i, name in enumerate(available_input_features, start=1)]))

    answer = input(
        "\nSelect one or several input features (indexes/ranges/names), [all], or [auto]: "
    ).strip()

    use_auto_feature_selection = answer.lower() == "auto"
    if use_auto_feature_selection:
        auto_pool_answer = input(
            "Auto mode candidate pool [all non-output CSV inputs] (use ! to exclude, e.g. !1 !3-5): "
        ).strip()
        try:
            selected_input_features = parse_multi_select(
                auto_pool_answer,
                available_input_features,
                allow_all=True,
                one_based=True,
            )
        except ValueError as err:
            print(f"Invalid auto candidate selection: {err}")
            exit()
    else:
        try:
            selected_input_features = parse_multi_select(
                answer,
                available_input_features,
                allow_all=True,
                one_based=True,
            )
        except ValueError as err:
            print(f"Invalid input feature selection: {err}")
            exit()

    resolved_inputs = (
        available_input_features if selected_input_features is None else selected_input_features
    )
    resolved_inputs, _ = remove_leakage_inputs_for_training(
        resolved_inputs,
        output_features=selected_output_features,
        context="raw_dataset_selection",
    )
    if not resolved_inputs:
        print("No input columns remain after automatic target-leakage removal.")
        exit()
    optional_future_quality_inputs = optional_input_overrides_for_selection(
        resolved_inputs, selected_output_features
    )
    validate_model_inputs(
        resolved_inputs,
        output_features=selected_output_features,
        optional_future_quality_inputs=optional_future_quality_inputs,
    )
    (
        sequential_features,
        sequential_groups,
        add_differences,
        difference_order,
        add_ratio_features,
        add_acceleration_features,
        add_normalized_change_features,
        create_rnn_windows,
        rnn_window_size,
        add_rolling_dynamics,
        rolling_window,
    ) = prompt_temporal_options(resolved_inputs)

    sync_prep_data_with_dataset(
        dataset_path=dataset_path,
        prep_folder=prep_folder,
        input_features=resolved_inputs,
        output_feature=selected_output_features,
        sequential_features=sequential_features,
        sequential_groups=sequential_groups,
        add_differences=add_differences,
        difference_order=difference_order,
        add_ratio_features=add_ratio_features,
        add_acceleration_features=add_acceleration_features,
        add_normalized_change_features=add_normalized_change_features,
        create_rnn_windows=create_rnn_windows,
        rnn_window_size=rnn_window_size,
        add_rolling_dynamics=add_rolling_dynamics,
        rolling_window=rolling_window,
        optional_future_quality_inputs=optional_future_quality_inputs,
    )

    X_train, X_test, y_train, y_test = read_prep_data(
        inputs=resolved_inputs,
        prep_folder=prep_folder,
        optional_future_quality_inputs=optional_future_quality_inputs,
    )
    _ = (X_test, y_test)
    return X_train, y_train.columns.tolist(), use_auto_feature_selection, reused_prep_data

def ask_with_default(prompt, default):
    answer = input(f"{prompt} [{default}]:").strip()
    return answer if answer else str(default)


def parse_epochs_input(answer, default_epochs):
    """Parse epochs input and detect optional CV early-stop mode."""
    text = answer.strip().lower()
    if text == "0":
        return [], False

    use_cv_early_stop = False
    parsed_epochs = []
    for token in answer.split():
        lowered = token.lower()
        if lowered in ("cv", "auto"):
            use_cv_early_stop = True
            continue
        if token.isdigit() and int(token) > 0:
            parsed_epochs.append(int(token))

    if not parsed_epochs and not use_cv_early_stop:
        parsed_epochs = list(default_epochs)
    return sorted(set(parsed_epochs)), use_cv_early_stop



def print_ann_training_robustness_notes():
    """Explain the safeguards used for neural-network training."""
    print("\nANN industrial robustness safeguards:")
    print(
        f"- learning_rate={learning_rate}: conservative AdamW step size reduces "
        "overshoot and run-to-run volatility on standardized tabular data; "
        f"ReduceLROnPlateau halves it after {lr_plateau_patience} flat checks "
        f"down to {min_learning_rate}."
    )
    print(
        f"- input_learning_rate={input_learning_rate}: input optimization moves "
        "data directly, so a smaller step prevents unrealistic optimized inputs."
    )
    print(
        f"- validation_fraction={validation_fraction}, patience={early_stopping_patience}: "
        "early stopping monitors held-out training data and restores the best weights, "
        "which limits memorization on the 755-sample dataset."
    )
    print(
        f"- ann_weight_decay={ann_weight_decay}: L2 regularization discourages oversized "
        "weights and improves generalization without adding model complexity."
    )
    print(
        "- BatchNorm1d after each hidden Linear/RBF basis layer stabilizes hidden "
        "activation scale; Dropout(p=0.10) reduces co-adaptation and small-sample "
        "memorization while preserving the existing model constructor API."
    )
    print(
        "- Xavier uniform initialization is applied to Linear layers so initial "
        "signal variance is balanced across input/output fan sizes before AdamW "
        "and the learning-rate scheduler begin updates."
    )
    print(
        f"- max_gradient_norm={max_gradient_norm}, max_input_delta={max_input_delta}: "
        "gradient and input-update bounds prevent unstable convergence and non-physical "
        "input drift."
    )
    print(
        f"- hidden size candidates={list_hidden_sizes}: for 755 samples, the current "
        "best [32, 16] is kept. One-hidden-layer candidates [16], [32], and "
        "[64] support the shallow models/RBFN; [16, 8], [48, 24], [64, 32], "
        "and [64, 32, 16] test moderate under/over-fitting around the current best."
    )

def infer_selected_features_from_table(table, fallback_features):
    """Infer selected features from the best-R2 row in a feature-selection table."""
    if table is None or table.empty:
        return list(fallback_features)

    if "R2" not in table.columns or "features" not in table.columns:
        return list(fallback_features)

    ranked = table.dropna(subset=["R2"])
    if ranked.empty:
        return list(fallback_features)

    best_row = ranked.loc[ranked["R2"].astype(float).idxmax()]
    features_text = str(best_row["features"]).strip()
    if not features_text:
        return list(fallback_features)

    selected = [f.strip() for f in features_text.split(",") if f.strip()]
    return selected if selected else list(fallback_features)


def compute_regression_report_metrics(y_true, y_pred):
    """
    Compute report-friendly regression metrics for the selected/best model.
    Metrics follow common WEKA-style definitions for RAE and RRSE.
    """
    y_true_arr = np.asarray(y_true, dtype=float)
    y_pred_arr = np.asarray(y_pred, dtype=float)

    if y_true_arr.ndim == 1:
        y_true_arr = y_true_arr.reshape(-1, 1)
    if y_pred_arr.ndim == 1:
        y_pred_arr = y_pred_arr.reshape(-1, 1)

    y_true_flat = y_true_arr.reshape(-1)
    y_pred_flat = y_pred_arr.reshape(-1)

    if y_true_flat.size == 0 or y_pred_flat.size == 0:
        return None

    if y_true_flat.size > 1:
        correlation = float(np.corrcoef(y_true_flat, y_pred_flat)[0, 1])
    else:
        correlation = np.nan

    abs_errors = np.abs(y_pred_flat - y_true_flat)
    sq_errors = (y_pred_flat - y_true_flat) ** 2
    mae = float(np.mean(abs_errors))
    rmse = float(np.sqrt(np.mean(sq_errors)))

    true_mean = float(np.mean(y_true_flat))
    rae_denominator = float(np.sum(np.abs(y_true_flat - true_mean)))
    rrse_denominator = float(np.sum((y_true_flat - true_mean) ** 2))
    rae = float((np.sum(abs_errors) / rae_denominator) * 100.0) if rae_denominator != 0 else np.nan
    rrse = float(np.sqrt(np.sum(sq_errors) / rrse_denominator) * 100.0) if rrse_denominator != 0 else np.nan

    return {
        "Correlation coefficient": correlation,
        "Mean absolute error": mae,
        "Root mean squared error": rmse,
        "Relative absolute error": rae,
        "Root relative squared error": rrse,
        "Total Number of Instances": int(y_true_arr.shape[0]),
    }

# Neural-network normalization uses sklearn StandardScaler exclusively.
# Each scaler is fitted only on the active training partition and then reused
# for validation/test/inference data so held-out statistics never leak into the model.

# Function to apply model on data and generate predictions
# Return predictions, MSE and R-Squared



def as_2d_float_array(data, name):
    """Convert pandas/numpy-like regression data to a finite 2D float array."""
    values = data.values if hasattr(data, "values") else data
    arr = np.asarray(values, dtype=float)
    if arr.ndim == 1:
        arr = arr.reshape(-1, 1)
    if arr.ndim != 2:
        raise ValueError(f"{name} must be a 1D or 2D array; got shape {arr.shape}.")
    if arr.shape[0] == 0:
        raise ValueError(f"{name} must contain at least one row.")
    if not np.isfinite(arr).all():
        raise ValueError(f"{name} contains NaN or infinite values.")
    return arr


def as_feature_matrix(data, name):
    """Convert tabular feature data to a finite 2D float matrix."""
    arr = as_2d_float_array(data, name)
    if arr.shape[1] == 0:
        raise ValueError(f"{name} must contain at least one feature column.")
    return arr



def dataframe_from_tabular(data, name, columns=None, index=None):
    """Convert tabular data to a finite DataFrame that preserves feature names."""
    if isinstance(data, pd.DataFrame):
        dataframe = data.copy()
    elif isinstance(data, pd.Series):
        dataframe = data.to_frame(name=columns[0] if columns else (data.name or name))
    else:
        array = np.asarray(data, dtype=float)
        if array.ndim == 1:
            array = array.reshape(-1, 1)
        if array.ndim != 2:
            raise ValueError(f"{name} must be a 1D or 2D array; got shape {array.shape}.")
        if columns is None:
            columns = [f"{name}_{idx}" for idx in range(array.shape[1])]
        dataframe = pd.DataFrame(array, columns=columns, index=index)
    if columns is not None:
        dataframe = dataframe.reindex(columns=columns)
    dataframe = dataframe.apply(pd.to_numeric, errors="coerce")
    if dataframe.shape[0] == 0:
        raise ValueError(f"{name} must contain at least one row.")
    if dataframe.shape[1] == 0:
        raise ValueError(f"{name} must contain at least one column.")
    if not np.isfinite(dataframe.to_numpy(dtype=float)).all():
        raise ValueError(f"{name} contains NaN or infinite values.")
    return dataframe


def target_dataframe(data, name, columns=None):
    """Convert target values to a DataFrame for named StandardScaler fitting."""
    dataframe = dataframe_from_tabular(data, name, columns=columns)
    if columns is None and dataframe.shape[1] == 1:
        dataframe.columns = ["target"]
    return dataframe


def fit_transform_dataframe(scaler, dataframe):
    """Fit a scaler on a DataFrame and return a DataFrame, not an ndarray."""
    return pd.DataFrame(
        scaler.fit_transform(dataframe),
        columns=dataframe.columns,
        index=dataframe.index,
    )


def transform_dataframe(scaler, dataframe):
    """Transform a DataFrame and preserve scaler feature names in the result."""
    columns = getattr(scaler, "feature_names_in_", dataframe.columns)
    return pd.DataFrame(
        scaler.transform(dataframe),
        columns=columns,
        index=dataframe.index,
    )


def inverse_transform_dataframe(scaler, dataframe):
    """Inverse-transform scaled DataFrame values without losing column names."""
    columns = getattr(scaler, "feature_names_in_", dataframe.columns)
    return pd.DataFrame(
        scaler.inverse_transform(dataframe),
        columns=columns,
        index=dataframe.index,
    )

def safe_r2_score(y_true, y_pred):
    """Compute finite multi-output R², returning None when undefined."""
    y_true_arr = as_2d_float_array(y_true, "y_true")
    y_pred_arr = as_2d_float_array(y_pred, "y_pred")
    if y_true_arr.shape != y_pred_arr.shape:
        raise ValueError(
            f"y_true and y_pred must have matching shapes; got "
            f"{y_true_arr.shape} and {y_pred_arr.shape}."
        )
    if y_true_arr.shape[0] < 2:
        return None
    score = r2_score(y_true_arr, y_pred_arr, multioutput="uniform_average")
    return float(score) if np.isfinite(score) else None


def initialize_linear_weights(module):
    """Initialize ANN layers with Xavier scaling and stable normalization defaults."""
    if isinstance(module, nn.Linear):
        nn.init.xavier_uniform_(module.weight)
        if module.bias is not None:
            nn.init.zeros_(module.bias)
    elif isinstance(module, nn.BatchNorm1d):
        nn.init.ones_(module.weight)
        nn.init.zeros_(module.bias)


def make_torch_model(model_class, input_size, hidden_sizes, output_size):
    """Instantiate a torch model while preserving older constructors."""
    if model_class.__name__ == "GRNN":
        return model_class(input_size, output_size=output_size)
    try:
        return model_class(input_size, hidden_sizes, output_size=output_size)
    except TypeError:
        return model_class(input_size, hidden_sizes)


def validate_hidden_size_compatibility(model, hidden_sizes):
    """Check whether the requested hidden-size layout matches a model class."""
    hidden_layers = getattr(model, "hidden_layers", None)
    return hidden_layers is None or len(hidden_sizes) == len(hidden_layers)


def load_model_state_dict(model, checkpoint_path):
    """Safely restore model weights from a checkpoint path when compatible."""
    if not checkpoint_path:
        return False
    if not os.path.isfile(checkpoint_path):
        print(f"Checkpoint not found: {checkpoint_path}. Training from initialized weights.")
        return False
    try:
        try:
            state_dict = torch.load(checkpoint_path, map_location="cpu", weights_only=True)
        except TypeError:
            state_dict = torch.load(checkpoint_path, map_location="cpu")
        model.load_state_dict(state_dict)
    except (OSError, RuntimeError, ValueError) as err:
        print(f"Could not load checkpoint '{checkpoint_path}' ({err}). Training from initialized weights.")
        return False
    print(f"Loaded pretrained weights from: {checkpoint_path}")
    return True


def parse_optimization_scope(answer):
    """Return training optimization scope: weights, inputs, or both."""
    normalized = (answer or "").strip().lower()
    mapping = {
        "": "weights",
        "w": "weights",
        "weight": "weights",
        "weights": "weights",
        "i": "inputs",
        "input": "inputs",
        "inputs": "inputs",
        "b": "both",
        "both": "both",
        "wi": "both",
        "iw": "both",
    }
    if normalized not in mapping:
        raise ValueError("Invalid optimization scope. Use weights, inputs, or both.")
    return mapping[normalized]


def prompt_optimized_input_features(input_features):
    print("\nInput-optimization target features:")
    print_selection_guide()
    print("\n".join([str(i) + ")" + name for i, name in enumerate(input_features, start=1)]))
    answer = input(
        "Select input features to optimize (indexes/ranges) or [all] [all]: "
    ).strip()
    if not answer or answer.lower() == "all":
        return None
    try:
        selected = parse_multi_select(answer, input_features, allow_all=True, one_based=True)
    except ValueError as err:
        print(f"Invalid optimized-input feature selection: {err}")
        exit()
    if selected is None:
        return None
    return [input_features.index(name) for name in selected]


def fit_model(
    model,
    X_train,
    X_test,
    y_train,
    y_test,
    num_epochs,
    display_steps=False,
    run=0,
    optimization_scope="weights",
    optimize_feature_indexes=None,
    pretrained_state_dict_path=None,
):
    """Train one ANN on tabular data with train-only normalization.

    Returns ``(predictions_original_scale, mse_original_scale, r2, model,
    optimized_inputs_report)`` for every success/failure path so callers can
    safely unpack results.  Scalers are fitted only on the training subset used
    for optimization and are then reused for validation/test data.
    """
    set_model_seed(model_seed + run)
    try:
        X_train_df = dataframe_from_tabular(X_train, "X_train")
        X_test_df = dataframe_from_tabular(X_test, "X_test", columns=X_train_df.columns)
        y_train_df = target_dataframe(y_train, "y_train")
        y_test_df = target_dataframe(y_test, "y_test", columns=y_train_df.columns)
        X_train_values = X_train_df.to_numpy(dtype=float)
        X_test_values = X_test_df.to_numpy(dtype=float)
        y_train_values = y_train_df.to_numpy(dtype=float)
        y_test_values = y_test_df.to_numpy(dtype=float)
    except ValueError as err:
        print(f"Invalid training data: {err}")
        return None, None, None, model, None

    if X_train_values.shape[0] != y_train_values.shape[0]:
        print("X_train and y_train row counts do not match.")
        return None, None, None, model, None
    if X_test_values.shape[0] != y_test_values.shape[0]:
        print("X_test and y_test row counts do not match.")
        return None, None, None, model, None

    optimize_weights = optimization_scope in ("weights", "both")
    optimize_inputs = optimization_scope in ("inputs", "both")
    if not optimize_weights and not optimize_inputs:
        print(f"Invalid optimization scope: {optimization_scope}")
        return None, None, None, model, None

    if optimize_feature_indexes is not None:
        optimize_feature_indexes = [int(idx) for idx in optimize_feature_indexes]
        invalid_indexes = [idx for idx in optimize_feature_indexes if idx < 0 or idx >= X_train_values.shape[1]]
        if invalid_indexes:
            print(f"Invalid optimized-input feature indexes: {invalid_indexes}")
            return None, None, None, model, None

    use_validation = len(X_train_values) >= 10 and validation_fraction > 0
    if use_validation:
        x_fit, x_val, y_fit, y_val = train_test_split(
            X_train_df,
            y_train_df,
            test_size=validation_fraction,
            random_state=data_seed + run,
            shuffle=True,
        )
    else:
        x_fit, y_fit = X_train_df, y_train_df
        x_val, y_val = None, None

    x_scaler = StandardScaler()
    y_scaler = StandardScaler()
    try:
        X_fit_normalized_df = fit_transform_dataframe(x_scaler, x_fit)
        X_val_normalized_df = transform_dataframe(x_scaler, x_val) if x_val is not None else None
        X_test_normalized_df = transform_dataframe(x_scaler, X_test_df)
        y_fit_normalized_df = fit_transform_dataframe(y_scaler, y_fit)
        y_val_normalized_df = transform_dataframe(y_scaler, y_val) if y_val is not None else None
        X_fit_normalized = torch.as_tensor(X_fit_normalized_df.to_numpy(), dtype=torch.float32)
        X_val_normalized = (
            torch.as_tensor(X_val_normalized_df.to_numpy(), dtype=torch.float32)
            if X_val_normalized_df is not None
            else None
        )
        X_test_normalized = torch.as_tensor(X_test_normalized_df.to_numpy(), dtype=torch.float32)
        y_fit_normalized = torch.as_tensor(y_fit_normalized_df.to_numpy(), dtype=torch.float32)
        y_val_normalized = (
            torch.as_tensor(y_val_normalized_df.to_numpy(), dtype=torch.float32)
            if y_val_normalized_df is not None
            else None
        )
    except ValueError as err:
        print(f"Could not normalize data safely: {err}")
        return None, None, None, model, None

    model.input_scaler_ = x_scaler
    model.target_scaler_ = y_scaler

    model.apply(initialize_linear_weights)
    load_model_state_dict(model, pretrained_state_dict_path)

    criterion = nn.MSELoss()
    optimizer = (
        optim.AdamW(model.parameters(), lr=learning_rate, weight_decay=ann_weight_decay, eps=1e-8)
        if optimize_weights
        else None
    )
    original_train_inputs = X_fit_normalized.clone().detach()
    train_inputs = X_fit_normalized.clone().detach().requires_grad_(optimize_inputs)
    input_optimizer = optim.Adam([train_inputs], lr=input_learning_rate, eps=1e-8) if optimize_inputs else None
    weight_scheduler = (
        optim.lr_scheduler.ReduceLROnPlateau(
            optimizer,
            mode="min",
            factor=lr_plateau_factor,
            patience=lr_plateau_patience,
            min_lr=min_learning_rate,
        )
        if optimizer is not None
        else None
    )

    best_monitor_loss = float("inf")
    best_epoch = 0
    best_state_dict = copy.deepcopy(model.state_dict())
    best_train_inputs = train_inputs.detach().clone()
    epochs_without_improvement = 0

    for epoch in range(int(num_epochs)):
        model.train()
        if optimizer is not None:
            optimizer.zero_grad(set_to_none=True)
        if input_optimizer is not None:
            input_optimizer.zero_grad(set_to_none=True)

        outputs = model(train_inputs)
        if outputs.shape != y_fit_normalized.shape:
            print(
                f"Model output shape {tuple(outputs.shape)} does not match "
                f"target shape {tuple(y_fit_normalized.shape)}."
            )
            return None, None, None, model, None
        if not torch.isfinite(outputs).all():
            print(f"Non-finite outputs detected at epoch {epoch + 1}")
            return None, None, None, model, None

        loss = criterion(outputs, y_fit_normalized)
        if not torch.isfinite(loss):
            print(f"Non-finite loss detected at epoch {epoch + 1}")
            return None, None, None, model, None

        loss.backward()

        if optimize_inputs and optimize_feature_indexes is not None and train_inputs.grad is not None:
            mask = torch.zeros_like(train_inputs.grad)
            mask[:, optimize_feature_indexes] = 1.0
            train_inputs.grad.mul_(mask)

        if optimize_weights:
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=max_gradient_norm)
            optimizer.step()

        if optimize_inputs:
            torch.nn.utils.clip_grad_norm_([train_inputs], max_norm=max_gradient_norm)
            input_optimizer.step()
            with torch.no_grad():
                lower_bound = original_train_inputs - max_input_delta
                upper_bound = original_train_inputs + max_input_delta
                train_inputs.clamp_(min=lower_bound, max=upper_bound)

        model.eval()
        with torch.no_grad():
            if X_val_normalized is not None:
                monitor_outputs = model(X_val_normalized)
                monitor_loss = criterion(monitor_outputs, y_val_normalized).item()
            else:
                monitor_loss = loss.item()

        if not np.isfinite(monitor_loss):
            print(f"Non-finite validation loss detected at epoch {epoch + 1}")
            return None, None, None, model, None

        if weight_scheduler is not None:
            weight_scheduler.step(monitor_loss)

        if monitor_loss < best_monitor_loss - early_stopping_min_delta:
            best_monitor_loss = monitor_loss
            best_epoch = epoch + 1
            best_state_dict = copy.deepcopy(model.state_dict())
            best_train_inputs = train_inputs.detach().clone()
            epochs_without_improvement = 0
        else:
            epochs_without_improvement += 1

        if (epoch + 1) % 10 == 0 and display_steps:
            print(
                f"Epoch [{epoch + 1}/{num_epochs}], "
                f"Train Loss: {loss.item():.6f}, Monitor Loss: {monitor_loss:.6f}"
            )

        if epochs_without_improvement >= early_stopping_patience:
            if display_steps:
                print(
                    f"Early stopping at epoch {epoch + 1}; "
                    f"best epoch {best_epoch} with monitor loss {best_monitor_loss:.6f}"
                )
            break

    if best_epoch == 0:
        print("Training did not complete a finite epoch.")
        return None, None, None, model, None

    model.load_state_dict(best_state_dict)
    train_inputs = best_train_inputs
    model.best_epoch_ = best_epoch
    model.best_monitor_loss_ = best_monitor_loss

    model.eval()
    with torch.no_grad():
        predictions_norm = model(X_test_normalized)
        if not torch.isfinite(predictions_norm).all():
            print("Non-finite predictions detected")
            return None, None, None, model, None

    try:
        predictions_norm_df = pd.DataFrame(
            predictions_norm.detach().cpu().numpy(),
            columns=y_train_df.columns,
        )
        predictions_np = inverse_transform_dataframe(y_scaler, predictions_norm_df).to_numpy()
        predictions_np = as_2d_float_array(predictions_np, "predictions")
        mse = float(np.mean((predictions_np - y_test_values) ** 2))
        r2 = safe_r2_score(y_test_values, predictions_np)
    except ValueError as err:
        print(f"Could not compute prediction metrics: {err}")
        return None, None, None, model, None

    optimized_inputs_report = None
    if optimize_inputs:
        with torch.no_grad():
            delta = train_inputs.detach() - original_train_inputs
            optimized_inputs_report = {
                "original_inputs": original_train_inputs.detach().cpu().numpy(),
                "optimized_inputs": train_inputs.detach().cpu().numpy(),
                "delta_inputs": delta.cpu().numpy(),
                "mean_abs_delta_per_feature": np.mean(np.abs(delta.cpu().numpy()), axis=0),
            }
    return predictions_np, mse, r2, model, optimized_inputs_report



def sklearn_model_supports_multioutput(model):
    """Return whether an sklearn-style estimator can fit a 2D target directly."""
    estimator = model.estimator if isinstance(model, RandomizedSearchCV) else model
    if isinstance(estimator, ExtraTreesRegressor):
        return True
    if hasattr(estimator, "_get_tags"):
        return bool(estimator._get_tags().get("multioutput", False))
    return False


def fit_model_on_split(
    model_class,
    X_train_fold,
    X_test_fold,
    y_train_fold,
    y_test_fold,
    num_epochs,
    hidden_sizes,
    run=0,
    optimization_scope="weights",
    optimize_feature_indexes=None,
):
    """Fit either an NN or sklearn model on one explicit train/test split."""
    X_train_fold, X_test_fold, _dropped_missing_columns = apply_missing_value_pipeline(
        X_train_fold,
        X_test_fold,
        verbose=False,
    )
    if is_torch_model(model_class):
        input_size = X_train_fold.shape[1]
        output_size = as_2d_float_array(y_train_fold, "y_train_fold").shape[1]
        model = make_torch_model(model_class, input_size, hidden_sizes, output_size)
        if not validate_hidden_size_compatibility(model, hidden_sizes):
            raise ValueError(f"Incompatible hidden sizes {hidden_sizes} for {model_class.__name__}.")
        predictions, mse, r2, _model, _optimized_inputs_report = fit_model(
            model,
            X_train_fold,
            X_test_fold,
            y_train_fold,
            y_test_fold,
            num_epochs,
            display_steps=False,
            run=run,
            optimization_scope=optimization_scope,
            optimize_feature_indexes=optimize_feature_indexes,
        )
        return predictions, mse, r2

    model = model_class(model_seed + run)
    predictions, mse, r2, _model = fit_sklearn_model(
        model, X_train_fold, X_test_fold, y_train_fold, y_test_fold
    )
    return predictions, mse, r2


def cross_validate_model(
    model_class,
    num_epochs,
    hidden_sizes,
    features=None,
    folds=MODEL_EVALUATION_CV_FOLDS,
    optimization_scope="weights",
    optimize_feature_indexes=None,
):
    """Run leakage-safe K-fold CV and report R²/RMSE mean and std."""
    X_train, X_test, y_train, y_test = read_prep_data(features)
    full_X = pd.concat([X_train, X_test], axis=0).reset_index(drop=True)
    full_y = pd.concat([y_train, y_test], axis=0).reset_index(drop=True)
    if len(full_X) < 2:
        return None

    n_splits = min(int(folds), len(full_X))
    if n_splits < 2:
        return None

    kfold = KFold(n_splits=n_splits, shuffle=True, random_state=data_seed)
    r2_scores = []
    rmse_scores = []
    for fold_index, (train_idx, val_idx) in enumerate(kfold.split(full_X), start=1):
        X_train_fold = full_X.iloc[train_idx]
        X_val_fold = full_X.iloc[val_idx]
        y_train_fold = full_y.iloc[train_idx]
        y_val_fold = full_y.iloc[val_idx]
        try:
            _predictions, mse, r2 = fit_model_on_split(
                model_class,
                X_train_fold,
                X_val_fold,
                y_train_fold,
                y_val_fold,
                num_epochs,
                hidden_sizes,
                run=fold_index,
                optimization_scope=optimization_scope,
                optimize_feature_indexes=optimize_feature_indexes,
            )
        except (TypeError, ValueError, RuntimeError) as err:
            print(f"CV fold {fold_index} failed: {err}")
            continue

        if r2 is None or mse is None:
            continue
        r2_scores.append(float(r2) * 100.0)
        rmse_scores.append(float(np.sqrt(mse)))

    if not r2_scores:
        return None

    return {
        "folds": len(r2_scores),
        "mean_r2": float(np.mean(r2_scores)),
        "std_r2": float(np.std(r2_scores)),
        "mean_rmse": float(np.mean(rmse_scores)),
        "std_rmse": float(np.std(rmse_scores)),
        "r2_scores": r2_scores,
        "rmse_scores": rmse_scores,
    }

def fit_sklearn_model(model, X_train, X_test, y_train, y_test):
    """Fit a scikit-learn regressor with train-only feature scaling."""
    X_train_df = dataframe_from_tabular(X_train, "X_train")
    X_test_df = dataframe_from_tabular(X_test, "X_test", columns=X_train_df.columns)
    y_train_values = as_2d_float_array(y_train, "y_train")
    y_test_values = as_2d_float_array(y_test, "y_test")

    scaler_x = StandardScaler()
    X_train_scaled_df = fit_transform_dataframe(scaler_x, X_train_df)
    X_test_scaled_df = transform_dataframe(scaler_x, X_test_df)
    X_train_scaled = X_train_scaled_df.to_numpy()
    X_test_scaled = X_test_scaled_df.to_numpy()

    is_multi_output = y_train_values.shape[1] > 1
    y_train_fit = y_train_values if is_multi_output else y_train_values.ravel()

    if isinstance(model, RandomizedSearchCV):
        search_cv_folds = min(int(model.cv), X_train_scaled.shape[0])
        if search_cv_folds < 2:
            print("Not enough training rows for ExtraTrees RandomizedSearchCV; fitting base estimator.")
            model = model.estimator
        else:
            model.cv = search_cv_folds

    fit_estimator = model
    if is_multi_output:
        model_name = getattr(model, "__class__", type(model)).__name__
        supports_multioutput = sklearn_model_supports_multioutput(model)
        if not supports_multioutput:
            print(
                f"{model_name} does not natively support multi-output regression. "
                "Wrapping it with MultiOutputRegressor."
            )
            fit_estimator = MultiOutputRegressor(model)

    fit_estimator.fit(X_train_scaled, y_train_fit)
    predictions = fit_estimator.predict(X_test_scaled)
    predictions_2d = as_2d_float_array(predictions, "predictions")

    if predictions_2d.shape != y_test_values.shape:
        raise ValueError(
            f"Prediction shape {predictions_2d.shape} does not match target shape {y_test_values.shape}."
        )

    fit_estimator.feature_scaler_ = scaler_x
    mse = float(np.mean((predictions_2d - y_test_values) ** 2))
    r2 = safe_r2_score(y_test_values, predictions_2d)
    return predictions_2d, mse, r2, fit_estimator


MODEL_COMPARISON_MODELS = (
    "RandomForest",
    "ExtraTrees",
    "XGBoost",
    "LightGBM",
    "CatBoost",
    "GradientBoosting",
    "ANN",
)
MODEL_COMPARISON_CV_FOLDS = 5
MODEL_COMPARISON_ANN_EPOCHS = 100
MODEL_COMPARISON_ANN_HIDDEN_SIZES = [32, 16]


def is_torch_model(model_class):
    return isinstance(model_class, type) and issubclass(model_class, nn.Module)


def _safe_mape_percent(y_true, y_pred, epsilon=1e-8):
    """Return MAPE (%) while ignoring zero/near-zero actual values."""
    y_true_arr = np.asarray(y_true, dtype=float)
    y_pred_arr = np.asarray(y_pred, dtype=float)
    denominator_mask = np.abs(y_true_arr) > epsilon
    if not denominator_mask.any():
        return np.nan
    return float(
        np.mean(
            np.abs((y_true_arr[denominator_mask] - y_pred_arr[denominator_mask]) / y_true_arr[denominator_mask])
        ) * 100.0
    )


def _target_regression_metrics(y_true, y_pred):
    """Compute R2, RMSE, MAE, and MAPE for a single target vector."""
    y_true_arr = np.asarray(y_true, dtype=float).reshape(-1)
    y_pred_arr = np.asarray(y_pred, dtype=float).reshape(-1)
    if y_true_arr.shape != y_pred_arr.shape:
        raise ValueError(f"Metric input shapes do not match: {y_true_arr.shape} and {y_pred_arr.shape}.")

    r2 = np.nan
    if y_true_arr.size >= 2:
        candidate_r2 = r2_score(y_true_arr, y_pred_arr)
        r2 = float(candidate_r2) if np.isfinite(candidate_r2) else np.nan

    errors = y_pred_arr - y_true_arr
    return {
        "R2": r2,
        "RMSE": float(np.sqrt(np.mean(errors ** 2))),
        "MAE": float(np.mean(np.abs(errors))),
        "MAPE": _safe_mape_percent(y_true_arr, y_pred_arr),
    }


def _build_model_comparison_factories(seed):
    """Build the exact model set requested for the cross-validation comparison."""
    factories = OrderedDict()
    factories["RandomForest"] = lambda fold_seed: RandomForestRegressor(
        n_estimators=300,
        random_state=fold_seed,
        n_jobs=-1,
    )
    factories["ExtraTrees"] = lambda fold_seed: ExtraTreesRegressor(
        n_estimators=300,
        random_state=fold_seed,
        n_jobs=-1,
    )

    if XGBRegressor is not None:
        factories["XGBoost"] = lambda fold_seed: XGBRegressor(
            n_estimators=300,
            random_state=fold_seed,
            objective="reg:squarederror",
            n_jobs=-1,
        )
    if LGBMRegressor is not None:
        factories["LightGBM"] = lambda fold_seed: LGBMRegressor(
            n_estimators=300,
            random_state=fold_seed,
            verbose=-1,
        )
    if CatBoostRegressor is not None:
        factories["CatBoost"] = lambda fold_seed: CatBoostRegressor(
            iterations=1200,
            learning_rate=0.03,
            depth=6,
            l2_leaf_reg=5.0,
            loss_function="RMSE",
            random_seed=fold_seed,
            verbose=False,
            allow_writing_files=False,
        )

    factories["GradientBoosting"] = lambda fold_seed: GradientBoostingRegressor(random_state=fold_seed)

    missing_models = [model_name for model_name in MODEL_COMPARISON_MODELS if model_name not in factories and model_name != "ANN"]
    if missing_models:
        print("Model comparison optional dependencies missing; skipped: " + ", ".join(missing_models))

    return factories


def _fit_comparison_model(model_name, model_factory, X_train_fold, X_val_fold, y_train_fold, y_val_fold, fold_seed):
    """Fit one comparison model and return validation predictions."""
    if model_name == "ANN":
        ann_model_class = globals().get("ReluFFNN") or globals().get("FFNN")
        if ann_model_class is None:
            raise ValueError("No ANN model class was found in models.py.")
        predictions, _mse, _r2 = fit_model_on_split(
            ann_model_class,
            X_train_fold,
            X_val_fold,
            y_train_fold,
            y_val_fold,
            MODEL_COMPARISON_ANN_EPOCHS,
            MODEL_COMPARISON_ANN_HIDDEN_SIZES,
            run=fold_seed,
            optimization_scope="weights",
        )
        if predictions is None:
            raise ValueError("ANN training did not produce validation predictions.")
        return predictions

    model = model_factory(fold_seed)
    predictions, _mse, _r2, _fit_estimator = fit_sklearn_model(
        model,
        X_train_fold,
        X_val_fold,
        y_train_fold,
        y_val_fold,
    )
    return predictions


def _summarize_model_target_folds(fold_metric_rows):
    """Average fold metrics for each target/model pair and sort by highest R2 per target."""
    fold_df = pd.DataFrame(fold_metric_rows)
    if fold_df.empty:
        return pd.DataFrame(), fold_df

    summary = (
        fold_df.groupby(["Target", "Model"], as_index=False)
        .agg(
            R2=("R2", "mean"),
            RMSE=("RMSE", "mean"),
            MAE=("MAE", "mean"),
            MAPE=("MAPE", "mean"),
            R2_std=("R2", "std"),
            RMSE_std=("RMSE", "std"),
            MAE_std=("MAE", "std"),
            MAPE_std=("MAPE", "std"),
            CV_folds=("Fold", "nunique"),
        )
        .sort_values(["Target", "R2"], ascending=[True, False], na_position="last")
        .reset_index(drop=True)
    )
    summary.insert(0, "Rank", summary.groupby("Target").cumcount() + 1)
    summary["Best for target"] = summary["Rank"].eq(1)
    ordered_columns = [
        "Target",
        "Rank",
        "Best for target",
        "Model",
        "R2",
        "RMSE",
        "MAE",
        "MAPE",
        "R2_std",
        "RMSE_std",
        "MAE_std",
        "MAPE_std",
        "CV_folds",
    ]
    return summary[ordered_columns], fold_df


def _write_model_comparison_workbook(summary_df, fold_df, report_path):
    """Write the model-comparison Excel workbook and highlight each target winner."""
    os.makedirs(os.path.dirname(report_path), exist_ok=True)

    best_df = pd.DataFrame()
    overall_df = pd.DataFrame()
    if not summary_df.empty:
        best_df = summary_df[summary_df["Best for target"]].copy()
        overall_df = (
            summary_df.groupby("Model", as_index=False)
            .agg(
                R2=("R2", "mean"),
                RMSE=("RMSE", "mean"),
                MAE=("MAE", "mean"),
                MAPE=("MAPE", "mean"),
                Targets=("Target", "nunique"),
            )
            .sort_values("R2", ascending=False, na_position="last")
            .reset_index(drop=True)
        )
        overall_df.insert(0, "Overall Rank", np.arange(1, len(overall_df) + 1))

    metadata_df = pd.DataFrame(
        [
            {
                "created_at_utc": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
                "cv_folds_requested": MODEL_COMPARISON_CV_FOLDS,
                "sort_rule": "Within each target, models are sorted by highest mean CV R2.",
                "best_model_rule": "Best for target = rank 1 by mean CV R2.",
                "ann_model": (globals().get("ReluFFNN") or globals().get("FFNN") or object).__name__,
                "ann_epochs": MODEL_COMPARISON_ANN_EPOCHS,
                "ann_hidden_sizes": str(MODEL_COMPARISON_ANN_HIDDEN_SIZES),
            }
        ]
    )

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Model Comparison", index=False)
        best_df.to_excel(writer, sheet_name="Best by Target", index=False)
        overall_df.to_excel(writer, sheet_name="Overall Average", index=False)
        fold_df.to_excel(writer, sheet_name="Fold Metrics", index=False)
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False)

        workbook = writer.book
        best_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        best_font = Font(bold=True, color="006100")
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 45)

        comparison_sheet = workbook["Model Comparison"]
        header_lookup = {cell.value: cell.column for cell in comparison_sheet[1]}
        best_column = header_lookup.get("Best for target")
        if best_column is not None:
            for row in range(2, comparison_sheet.max_row + 1):
                if comparison_sheet.cell(row=row, column=best_column).value is True:
                    for column in range(1, comparison_sheet.max_column + 1):
                        comparison_sheet.cell(row=row, column=column).fill = best_fill
                        comparison_sheet.cell(row=row, column=column).font = best_font

    print(f"Saved model comparison report: {report_path}")


def generate_model_comparison_report(X_train, X_test, y_train, y_test, report_path=MODEL_COMPARISON_REPORT_PATH):
    """Compare requested regressors with 5-fold CV and save reports/model_comparison.xlsx."""
    full_X = pd.concat([X_train, X_test], axis=0).reset_index(drop=True)
    full_y = pd.concat([y_train, y_test], axis=0).reset_index(drop=True)
    if len(full_X) < 2:
        print("Model comparison skipped: at least two rows are required for cross-validation.")
        return None

    n_splits = min(MODEL_COMPARISON_CV_FOLDS, len(full_X))
    if n_splits < MODEL_COMPARISON_CV_FOLDS:
        print(f"Model comparison warning: using {n_splits} folds because only {len(full_X)} rows are available.")
    kfold = KFold(n_splits=n_splits, shuffle=True, random_state=data_seed)
    comparison_factories = _build_model_comparison_factories(model_seed)
    comparison_factories["ANN"] = None

    fold_metric_rows = []
    target_names = list(full_y.columns)
    print("\n================= Model Comparison (5-fold CV) =================")
    for model_name in MODEL_COMPARISON_MODELS:
        if model_name not in comparison_factories:
            continue
        print(f"Evaluating {model_name}...")
        for fold_index, (train_idx, val_idx) in enumerate(kfold.split(full_X), start=1):
            X_train_fold = full_X.iloc[train_idx]
            X_val_fold = full_X.iloc[val_idx]
            y_train_fold = full_y.iloc[train_idx]
            y_val_fold = full_y.iloc[val_idx]
            fold_seed = model_seed + fold_index
            try:
                predictions = _fit_comparison_model(
                    model_name,
                    comparison_factories[model_name],
                    X_train_fold,
                    X_val_fold,
                    y_train_fold,
                    y_val_fold,
                    fold_seed,
                )
            except (TypeError, ValueError, RuntimeError) as err:
                print(f"{model_name} fold {fold_index} skipped: {err}")
                continue

            try:
                predictions_2d = as_2d_float_array(predictions, f"{model_name} predictions")
                y_val_values = as_2d_float_array(y_val_fold, "y_val_fold")
            except ValueError as err:
                print(f"{model_name} fold {fold_index} skipped: {err}")
                continue
            for target_index, target_name in enumerate(target_names):
                metrics = _target_regression_metrics(
                    y_val_values[:, target_index],
                    predictions_2d[:, target_index],
                )
                fold_metric_rows.append(
                    {
                        "Target": target_name,
                        "Model": model_name,
                        "Fold": fold_index,
                        **metrics,
                    }
                )

    summary_df, fold_df = _summarize_model_target_folds(fold_metric_rows)
    if summary_df.empty:
        print("Model comparison skipped: no model produced valid fold metrics.")
        return None

    _write_model_comparison_workbook(summary_df, fold_df, report_path)
    print(tabulate(summary_df[["Target", "Rank", "Model", "R2", "RMSE", "MAE", "MAPE"]], headers="keys", tablefmt="github", showindex=False))
    print("===============================================================\n")
    return report_path



def select_epochs_with_cv_early_stopping(
    model_class,
    hidden_sizes,
    features=None,
    folds=5,
    patience=early_stopping_patience,
    min_delta=early_stopping_min_delta,
):
    """Recommend an epoch count using leakage-safe K-Fold early stopping."""
    X_train, _, y_train, _ = read_prep_data(features)
    full_X = X_train.reset_index(drop=True)
    full_y = y_train.reset_index(drop=True)
    if len(full_X) < 2:
        return None

    n_splits = min(int(folds), len(full_X))
    if n_splits < 2:
        return None

    kfold = KFold(n_splits=n_splits, shuffle=True, random_state=data_seed)
    best_epochs = []
    input_size = full_X.shape[1]
    output_size = as_2d_float_array(full_y, "full_y").shape[1]
    max_epochs = max(list_epochs) if list_epochs else 200

    for fold_index, (train_idx, val_idx) in enumerate(kfold.split(full_X), start=1):
        x_train_fold = full_X.iloc[train_idx]
        x_val_fold = full_X.iloc[val_idx]
        y_train_fold = full_y.iloc[train_idx]
        y_val_fold = full_y.iloc[val_idx]

        x_scaler = StandardScaler()
        y_scaler = StandardScaler()
        try:
            x_train_scaled = fit_transform_dataframe(x_scaler, x_train_fold)
            x_val_scaled = transform_dataframe(x_scaler, x_val_fold)
            y_train_frame = target_dataframe(y_train_fold, "y_train_fold")
            y_val_frame = target_dataframe(y_val_fold, "y_val_fold", columns=y_train_frame.columns)
            y_train_values = y_train_frame.to_numpy(dtype=float)
            y_val_values = y_val_frame.to_numpy(dtype=float)
            y_train_scaled = fit_transform_dataframe(y_scaler, y_train_frame)
            y_val_scaled = transform_dataframe(y_scaler, y_val_frame)
            x_train_tensor = torch.as_tensor(x_train_scaled.to_numpy(), dtype=torch.float32)
            x_val_tensor = torch.as_tensor(x_val_scaled.to_numpy(), dtype=torch.float32)
            y_train_tensor = torch.as_tensor(y_train_scaled.to_numpy(), dtype=torch.float32)
            y_val_tensor = torch.as_tensor(y_val_scaled.to_numpy(), dtype=torch.float32)
            model = make_torch_model(model_class, input_size, hidden_sizes, output_size)
        except (TypeError, ValueError) as err:
            print(f"CV fold {fold_index} skipped: {err}")
            continue

        if not validate_hidden_size_compatibility(model, hidden_sizes):
            return None

        set_model_seed(model_seed + fold_index)
        model.apply(initialize_linear_weights)
        criterion = nn.MSELoss()
        optimizer = optim.AdamW(model.parameters(), lr=learning_rate, weight_decay=ann_weight_decay, eps=1e-8)
        scheduler = optim.lr_scheduler.ReduceLROnPlateau(
            optimizer,
            mode="min",
            factor=lr_plateau_factor,
            patience=lr_plateau_patience,
            min_lr=min_learning_rate,
        )

        best_val_loss = float("inf")
        best_epoch = None
        epochs_without_improvement = 0

        for epoch in range(1, int(max_epochs) + 1):
            model.train()
            optimizer.zero_grad(set_to_none=True)
            train_outputs = model(x_train_tensor)
            train_loss = criterion(train_outputs, y_train_tensor)
            if not torch.isfinite(train_loss):
                break
            train_loss.backward()
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=max_gradient_norm)
            optimizer.step()

            model.eval()
            with torch.no_grad():
                val_outputs = model(x_val_tensor)
                val_loss = criterion(val_outputs, y_val_tensor).item()

            if not np.isfinite(val_loss):
                break
            scheduler.step(val_loss)

            if val_loss < best_val_loss - min_delta:
                best_val_loss = val_loss
                best_epoch = epoch
                epochs_without_improvement = 0
            else:
                epochs_without_improvement += 1

            if epochs_without_improvement >= patience:
                break

        if best_epoch is not None:
            best_epochs.append(best_epoch)

    if not best_epochs:
        return None
    return int(np.median(best_epochs))


############################ Feature Selection ####################
# Selects the best combination of features by removing features one by one
# Search about Backward Feature Elimination
# Sensitivity Analysis Functions



def train_single_model(model_class, num_epochs, hidden_sizes, features=None, run=0):
    """Train one model instance and return trained model with raw train/test splits."""
    X_train, X_test, y_train, y_test = read_prep_data(features)
    input_size = X_train.shape[1]
    output_size = y_train.shape[1] if hasattr(y_train, "shape") and len(y_train.shape) > 1 else 1
    model = make_torch_model(model_class, input_size, hidden_sizes, output_size)

    _, _, r2, model, _ = fit_model(
        model, X_train, X_test, y_train, y_test, num_epochs, display_steps=False, run=run
    )
    return model, X_train, X_test, y_train, y_test, r2


def _parse_positive_int_with_default(answer, default_value, name):
    """Parse a positive integer from interactive input while preserving defaults."""
    if not answer:
        return default_value
    try:
        value = int(answer)
    except ValueError:
        print(f"Invalid {name} '{answer}'. Using default value {default_value}.")
        return default_value
    if value <= 0:
        print(f"{name} must be > 0. Using default value {default_value}.")
        return default_value
    return value


def _train_model_for_shap(model_class, inputs, num_epochs, hidden_sizes):
    """Train the selected model type and return fitted model plus SHAP data splits."""
    if is_torch_model(model_class):
        model, X_train, X_test, _, _, r2 = train_single_model(
            model_class, num_epochs, hidden_sizes, features=inputs, run=0
        )
        if r2 is None:
            print("Could not train model for SHAP analysis.")
            return None, None, None, None, False
        scaler = getattr(model, "input_scaler_", None)
        if scaler is None:
            scaler = StandardScaler().fit(X_train)
        X_train_explain = transform_dataframe(scaler, dataframe_from_tabular(X_train, "X_train")).to_numpy()
        X_test_explain = transform_dataframe(
            scaler,
            dataframe_from_tabular(X_test, "X_test", columns=getattr(scaler, "feature_names_in_", X_train.columns)),
        ).to_numpy()
        return model, X_train, X_train_explain, X_test_explain, True

    X_train, X_test, y_train, y_test = read_prep_data(inputs)
    try:
        model = model_class(model_seed)
        _, _, r2, fitted_model = fit_sklearn_model(model, X_train, X_test, y_train, y_test)
    except (TypeError, ValueError, RuntimeError) as err:
        print(f"Could not train sklearn model for SHAP analysis: {err}")
        return None, None, None, None, False
    if r2 is None:
        print("Could not train sklearn model for SHAP analysis.")
        return None, None, None, None, False
    scaler = getattr(fitted_model, "feature_scaler_", None)
    if scaler is None:
        scaler = StandardScaler().fit(X_train)
    X_train_explain = transform_dataframe(scaler, dataframe_from_tabular(X_train, "X_train")).to_numpy()
    X_test_explain = transform_dataframe(
        scaler,
        dataframe_from_tabular(X_test, "X_test", columns=getattr(scaler, "feature_names_in_", X_train.columns)),
    ).to_numpy()
    return fitted_model, X_train, X_train_explain, X_test_explain, False


def _unwrap_tree_estimators_for_shap(model):
    """Return fitted tree estimators that SHAP TreeExplainer can explain directly."""
    if isinstance(model, MultiOutputRegressor):
        return list(model.estimators_)
    if isinstance(model, RandomizedSearchCV):
        return [model.best_estimator_]
    return [model]


def _is_supported_tree_model(model):
    """Detect tree models requested for fast SHAP support, including wrappers."""
    tree_type_names = {
        "ExtraTreesRegressor",
        "XGBRegressor",
        "LGBMRegressor",
        "CatBoostRegressor",
    }
    return all(type(estimator).__name__ in tree_type_names for estimator in _unwrap_tree_estimators_for_shap(model))


def _torch_predict_fn(model):
    def predict_fn(x):
        with torch.no_grad():
            x_tensor = torch.tensor(x, dtype=torch.float32)
            pred = model(x_tensor).detach().numpy()
        return pred

    return predict_fn


def _generic_predict_fn(model):
    def predict_fn(x):
        pred = model.predict(x)
        return as_2d_float_array(pred, "SHAP predictions")

    return predict_fn


def _compute_tree_shap_values(shap_module, model, explain_points):
    """Compute TreeExplainer SHAP values, preserving multi-output information."""
    shap_values_by_output = []
    for estimator in _unwrap_tree_estimators_for_shap(model):
        explainer = shap_module.TreeExplainer(estimator)
        shap_values_by_output.append(explainer.shap_values(explain_points))
    if len(shap_values_by_output) == 1:
        return shap_values_by_output[0]
    return shap_values_by_output


def _mean_abs_shap_by_feature(shap_values, n_features):
    """Collapse SHAP output/sample axes into one mean absolute value per feature."""
    if isinstance(shap_values, list):
        arrays = [np.asarray(values) for values in shap_values]
        feature_vectors = [_mean_abs_shap_by_feature(values, n_features) for values in arrays]
        return np.mean(np.vstack(feature_vectors), axis=0)

    shap_abs = np.abs(np.asarray(shap_values))
    if shap_abs.ndim == 1:
        return shap_abs.reshape(-1)

    feature_axes = [axis for axis, size in enumerate(shap_abs.shape) if size == n_features]
    if not feature_axes:
        return shap_abs.reshape(-1)

    feature_axis = feature_axes[-1]
    shap_abs = np.moveaxis(shap_abs, feature_axis, -1)
    return shap_abs.reshape(-1, n_features).mean(axis=0)


def _save_shap_plots(shap_module, shap_values, explain_points, feature_names, model_slug):
    """Save SHAP summary and bar plots for the current model."""
    os.makedirs("plots", exist_ok=True)
    summary_path = os.path.join("plots", f"shap_summary_{model_slug}.png")
    bar_path = os.path.join("plots", f"shap_bar_{model_slug}.png")

    plot_values = shap_values
    if isinstance(shap_values, list) and shap_values:
        plot_values = shap_values[0]
        if len(shap_values) > 1:
            print("Multiple SHAP outputs detected; plots use the first output while CSV tables average all outputs.")
    else:
        plot_array = np.asarray(shap_values)
        if plot_array.ndim > 2:
            feature_axes = [axis for axis, size in enumerate(plot_array.shape) if size == len(feature_names)]
            if feature_axes:
                feature_axis = feature_axes[-1]
                plot_array = np.moveaxis(plot_array, feature_axis, 1)
                plot_values = plot_array.reshape(plot_array.shape[0], plot_array.shape[1], -1)[:, :, 0]
                print("Multiple SHAP outputs detected; plots use the first output while CSV tables average all outputs.")

    try:
        shap_module.summary_plot(plot_values, explain_points, feature_names=feature_names, show=False)
        plt.tight_layout()
        plt.savefig(summary_path, dpi=200, bbox_inches="tight")
        plt.close()
        print(f"SHAP summary plot saved at {summary_path}")
    except Exception as err:
        plt.close()
        print(f"Could not save SHAP summary plot: {err}")

    try:
        shap_module.summary_plot(
            plot_values,
            explain_points,
            feature_names=feature_names,
            plot_type="bar",
            show=False,
        )
        plt.tight_layout()
        plt.savefig(bar_path, dpi=200, bbox_inches="tight")
        plt.close()
        print(f"SHAP bar plot saved at {bar_path}")
    except Exception as err:
        plt.close()
        print(f"Could not save SHAP bar plot: {err}")

    # Preserve the historical filename for callers that expect this artifact.
    if model_slug != "best_model":
        legacy_summary_path = os.path.join("plots", "shap_summary.png")
        try:
            shap_module.summary_plot(plot_values, explain_points, feature_names=feature_names, show=False)
            plt.tight_layout()
            plt.savefig(legacy_summary_path, dpi=200, bbox_inches="tight")
            plt.close()
            print(f"SHAP summary plot saved at {legacy_summary_path}")
        except Exception:
            plt.close()


def _save_shap_tables(shap_table, model_slug):
    """Save ranked SHAP tables and the top-20 feature subset."""
    os.makedirs("tables", exist_ok=True)
    ranked_path = os.path.join("tables", f"shap_ranked_importance_{model_slug}.csv")
    top20_path = os.path.join("tables", f"shap_top20_features_{model_slug}.csv")
    legacy_ranked_path = os.path.join("tables", "shap_ranked_importance.csv")
    legacy_top20_path = os.path.join("tables", "shap_top20_features.csv")

    shap_table.to_csv(ranked_path, index=False)
    shap_table.head(20).to_csv(top20_path, index=False)
    shap_table.to_csv(legacy_ranked_path, index=False)
    shap_table.head(20).to_csv(legacy_top20_path, index=False)
    print(f"SHAP ranked importance table saved at {ranked_path}")
    print(f"SHAP top 20 features saved at {top20_path}")


SHAP_RESULTS_DIR = os.path.join("results", "shap")
SHAP_MAX_EXPLAIN_SAMPLES = 200


def _target_shap_tree_model_factories(seed):
    """Return target-level tree-model candidates for automatic SHAP analysis."""
    factories = [
        (
            "RandomForestRegressor",
            lambda: RandomForestRegressor(
                n_estimators=500,
                random_state=seed,
                n_jobs=-1,
            ),
        ),
        (
            "ExtraTreesRegressor",
            lambda: ExtraTreesRegressor(
                n_estimators=500,
                random_state=seed,
                n_jobs=-1,
            ),
        ),
        (
            "GradientBoostingRegressor",
            lambda: GradientBoostingRegressor(random_state=seed),
        ),
        (
            "AdaBoostRegressor",
            lambda: AdaBoostRegressor(random_state=seed),
        ),
        (
            "DecisionTreeRegressor",
            lambda: DecisionTreeRegressor(random_state=seed),
        ),
    ]

    if XGBRegressor is not None:
        factories.append(
            (
                "XGBoostRegressor",
                lambda: XGBRegressor(
                    n_estimators=500,
                    random_state=seed,
                    objective="reg:squarederror",
                ),
            )
        )
    if LGBMRegressor is not None:
        factories.append(
            (
                "LightGBMRegressor",
                lambda: LGBMRegressor(
                    n_estimators=500,
                    random_state=seed,
                    verbose=-1,
                ),
            )
        )
    if CatBoostRegressor is not None:
        factories.append(("CatBoostRegressor", lambda: make_catboost_regressor(seed)))

    return factories


def _prepare_target_shap_data(X_train, X_test, y_train, y_test, target):
    """Return finite numeric train/test matrices for one target."""
    X_train_target = X_train.copy().replace([np.inf, -np.inf], np.nan)
    X_test_target = X_test.copy().replace([np.inf, -np.inf], np.nan)
    y_train_target = pd.to_numeric(y_train[target], errors="coerce")
    y_test_target = pd.to_numeric(y_test[target], errors="coerce")

    numeric_train_parts = []
    numeric_test_parts = []
    kept_features = []
    for feature in X_train_target.columns:
        train_series = pd.to_numeric(X_train_target[feature], errors="coerce")
        test_series = pd.to_numeric(X_test_target[feature], errors="coerce")
        fill_value = train_series.median()
        if pd.isna(fill_value):
            continue
        numeric_train_parts.append(train_series.fillna(fill_value).astype(float).to_frame(feature))
        numeric_test_parts.append(test_series.fillna(fill_value).astype(float).to_frame(feature))
        kept_features.append(feature)

    if not numeric_train_parts or y_train_target.notna().sum() < 2:
        return None

    X_train_numeric = pd.concat(numeric_train_parts, axis=1)
    X_test_numeric = pd.concat(numeric_test_parts, axis=1)
    train_mask = y_train_target.notna()
    test_mask = y_test_target.notna()

    X_train_numeric = X_train_numeric.loc[train_mask]
    y_train_target = y_train_target.loc[train_mask].astype(float)
    X_test_numeric = X_test_numeric.loc[test_mask]
    y_test_target = y_test_target.loc[test_mask].astype(float)

    constant_features = [feature for feature in kept_features if X_train_numeric[feature].nunique(dropna=False) <= 1]
    if constant_features:
        X_train_numeric = X_train_numeric.drop(columns=constant_features)
        X_test_numeric = X_test_numeric.drop(columns=constant_features)

    if X_train_numeric.empty or X_test_numeric.empty or y_test_target.empty:
        return None

    return X_train_numeric, X_test_numeric, y_train_target, y_test_target


def _fit_best_tree_model_for_target(X_train, X_test, y_train, y_test, target, seed=None):
    """Train tree candidates for one target and return the best fitted model by test R²."""
    if seed is None:
        seed = model_seed

    prepared = _prepare_target_shap_data(X_train, X_test, y_train, y_test, target)
    if prepared is None:
        print(f"Skipping SHAP for {target}: not enough finite numeric rows/features.")
        return None

    X_train_target, X_test_target, y_train_target, y_test_target = prepared
    best = None
    rows = []
    for model_name, factory in _target_shap_tree_model_factories(seed):
        try:
            model = factory()
            model.fit(X_train_target, y_train_target)
            predictions = model.predict(X_test_target)
            score = safe_r2_score(y_test_target, predictions)
            score_for_sort = -np.inf if score is None else score
            rows.append({"Model": model_name, "R2": score})
            if best is None or score_for_sort > best["score_for_sort"]:
                best = {
                    "model_name": model_name,
                    "model": model,
                    "score": score,
                    "score_for_sort": score_for_sort,
                    "X_train": X_train_target,
                    "X_test": X_test_target,
                    "y_train": y_train_target,
                    "y_test": y_test_target,
                }
        except Exception as err:
            print(f"SHAP tree candidate skipped for {target} ({model_name}): {err}")

    if best is None:
        print(f"Skipping SHAP for {target}: no tree model could be trained.")
        return None

    metrics_table = pd.DataFrame(rows).sort_values(by="R2", ascending=False, na_position="last")
    target_dir = os.path.join(SHAP_RESULTS_DIR, _safe_report_filename_part(target))
    os.makedirs(target_dir, exist_ok=True)
    metrics_table.to_csv(os.path.join(target_dir, "tree_model_scores.csv"), index=False)
    score_text = "undefined" if best["score"] is None else f"{best['score']:.4f}"
    print(f"Best tree model for {target}: {best['model_name']} (test R2={score_text})")
    return best


def _coerce_shap_values_for_single_target(shap_values, n_features):
    """Return a 2D SHAP matrix for a single target and feature count."""
    if isinstance(shap_values, list):
        shap_values = shap_values[0]
    shap_array = np.asarray(shap_values)
    if shap_array.ndim == 3:
        feature_axes = [axis for axis, size in enumerate(shap_array.shape) if size == n_features]
        if feature_axes:
            feature_axis = feature_axes[-1]
            shap_array = np.moveaxis(shap_array, feature_axis, 1)
            shap_array = shap_array[:, :, 0]
    if shap_array.ndim != 2:
        shap_array = shap_array.reshape(shap_array.shape[0], -1)
    return shap_array


def _save_target_shap_plots(shap_module, shap_values, explain_points, feature_names, output_dir):
    """Save summary, bar, and dependence SHAP plots for one target."""
    os.makedirs(output_dir, exist_ok=True)
    summary_path = os.path.join(output_dir, "shap_summary.png")
    bar_path = os.path.join(output_dir, "shap_bar.png")
    dependence_path = os.path.join(output_dir, "shap_dependence.png")

    top_feature_index = int(np.argmax(np.abs(shap_values).mean(axis=0)))
    top_feature_name = feature_names[top_feature_index]

    shap_module.summary_plot(shap_values, explain_points, feature_names=feature_names, show=False)
    plt.tight_layout()
    plt.savefig(summary_path, dpi=200, bbox_inches="tight")
    plt.close()

    shap_module.summary_plot(
        shap_values,
        explain_points,
        feature_names=feature_names,
        plot_type="bar",
        show=False,
    )
    plt.tight_layout()
    plt.savefig(bar_path, dpi=200, bbox_inches="tight")
    plt.close()

    shap_module.dependence_plot(
        top_feature_index,
        shap_values,
        explain_points,
        feature_names=feature_names,
        show=False,
    )
    plt.tight_layout()
    plt.savefig(dependence_path, dpi=200, bbox_inches="tight")
    plt.close()

    return {
        "summary": summary_path,
        "bar": bar_path,
        "dependence": dependence_path,
        "dependence_feature": top_feature_name,
    }


def generate_target_shap_analysis(X_train, X_test, y_train, y_test, targets=None, results_dir=SHAP_RESULTS_DIR):
    """Train the best tree model per target, save SHAP plots, and print top refinery variables."""
    import importlib.util
    shap_spec = importlib.util.find_spec("shap")
    if shap_spec is None:
        print("SHAP target analysis skipped: shap is not installed. Install it with: pip install shap")
        return []
    import shap

    targets = list(targets or y_train.columns)
    os.makedirs(results_dir, exist_ok=True)
    generated = []
    print("\n================= Target SHAP Analysis =================")
    for target in targets:
        if target not in y_train.columns or target not in y_test.columns:
            print(f"Skipping SHAP for {target}: target is not present in train/test targets.")
            continue

        print(f"\nCalculating SHAP analysis for target: {target}")
        best = _fit_best_tree_model_for_target(X_train, X_test, y_train, y_test, target)
        if best is None:
            continue

        X_explain = best["X_test"].head(min(SHAP_MAX_EXPLAIN_SAMPLES, len(best["X_test"])))
        feature_names = list(X_explain.columns)
        try:
            explainer = shap.TreeExplainer(best["model"])
            shap_values = explainer.shap_values(X_explain)
            shap_values = _coerce_shap_values_for_single_target(shap_values, len(feature_names))
        except Exception as err:
            print(f"Skipping SHAP plots for {target}: could not calculate TreeExplainer values ({err}).")
            continue

        if shap_values.shape[1] != len(feature_names):
            print(
                f"Skipping SHAP plots for {target}: got {shap_values.shape[1]} SHAP features "
                f"for {len(feature_names)} input features."
            )
            continue

        mean_abs_shap = np.abs(shap_values).mean(axis=0)
        total_importance = float(mean_abs_shap.sum())
        normalized = mean_abs_shap / total_importance if total_importance > 0 else np.zeros_like(mean_abs_shap)
        top_table = pd.DataFrame(
            {
                "Rank": np.arange(1, len(feature_names) + 1),
                "Variable": feature_names,
                "Mean(|SHAP|)": mean_abs_shap,
                "Normalized Importance": normalized,
            }
        ).sort_values(by="Mean(|SHAP|)", ascending=False).reset_index(drop=True)
        top_table["Rank"] = np.arange(1, len(top_table) + 1)
        top_10 = top_table.head(10)

        target_dir = os.path.join(results_dir, _safe_report_filename_part(target))
        try:
            plot_paths = _save_target_shap_plots(shap, shap_values, X_explain, feature_names, target_dir)
        except Exception as err:
            plt.close()
            print(f"SHAP plot save failed for {target}: {err}")
            plot_paths = {}

        top_table.to_csv(os.path.join(target_dir, "shap_ranked_variables.csv"), index=False)
        top_10.to_csv(os.path.join(target_dir, "top10_refinery_variables.csv"), index=False)
        generated.append(
            {
                "target": target,
                "model": best["model_name"],
                "r2": best["score"],
                "directory": target_dir,
                "plots": plot_paths,
            }
        )

        print(f"Saved SHAP outputs for {target}: {target_dir}")
        if plot_paths:
            print(f" - shap_summary.png: {plot_paths['summary']}")
            print(f" - shap_bar.png: {plot_paths['bar']}")
            print(
                f" - shap_dependence.png: {plot_paths['dependence']} "
                f"(feature: {plot_paths['dependence_feature']})"
            )
        print(f"Top 10 most influential refinery variables for {target}:")
        print(tabulate(top_10, headers="keys", tablefmt="github", showindex=False))
    print("========================================================\n")
    return generated


def shap_feature_importance(model_class, inputs, num_epochs, hidden_sizes, model_name=None):
    """Compute and save SHAP feature importance for neural nets and supported tree models."""
    import importlib.util
    shap_spec = importlib.util.find_spec("shap")
    if shap_spec is None:
        print("SHAP is not installed. Install it with: pip install shap")
        return None
    import shap

    model, X_train, X_train_explain, X_test_explain, is_nn_model = _train_model_for_shap(
        model_class, inputs, num_epochs, hidden_sizes
    )
    if model is None:
        return None

    background_default = min(25, len(X_train_explain))
    explain_default = min(20, len(X_test_explain))
    nsamples_default = 100

    background_size = _parse_positive_int_with_default(
        input(f"Background sample size [{background_default}]:").strip(),
        background_default,
        "background sample size",
    )
    explain_size = _parse_positive_int_with_default(
        input(f"Number of test samples to explain [{explain_default}]:").strip(),
        explain_default,
        "number of test samples",
    )

    background_size = max(1, min(background_size, len(X_train_explain)))
    explain_size = max(1, min(explain_size, len(X_test_explain)))
    background = X_train_explain[:background_size]
    explain_points = X_test_explain[:explain_size]

    print("Computing SHAP values. This may take some time ...")
    if is_nn_model:
        model.eval()
        nsamples = _parse_positive_int_with_default(
            input(f"Kernel SHAP nsamples [{nsamples_default}]:").strip(),
            nsamples_default,
            "Kernel SHAP nsamples",
        )
        explainer = shap.KernelExplainer(_torch_predict_fn(model), background)
        shap_values = explainer.shap_values(explain_points, nsamples=nsamples)
    elif _is_supported_tree_model(model):
        shap_values = _compute_tree_shap_values(shap, model, explain_points)
    else:
        nsamples = _parse_positive_int_with_default(
            input(f"Kernel SHAP nsamples [{nsamples_default}]:").strip(),
            nsamples_default,
            "Kernel SHAP nsamples",
        )
        explainer = shap.KernelExplainer(_generic_predict_fn(model), background)
        shap_values = explainer.shap_values(explain_points, nsamples=nsamples)

    feature_names = list(X_train.columns)
    mean_abs_shap = _mean_abs_shap_by_feature(shap_values, len(feature_names))
    mean_abs_shap = np.asarray(mean_abs_shap).reshape(-1)

    if mean_abs_shap.shape[0] != len(feature_names):
        print(
            "Could not build SHAP table because feature and SHAP dimensions do not match. "
            f"Got {mean_abs_shap.shape[0]} SHAP values for {len(feature_names)} features. "
            f"Raw SHAP shape: {np.asarray(shap_values, dtype=object).shape}"
        )
        return None

    total_importance = float(np.sum(mean_abs_shap))
    if total_importance > 0:
        normalized_importance = mean_abs_shap / total_importance
    else:
        normalized_importance = np.zeros_like(mean_abs_shap)

    shap_table = pd.DataFrame(
        {
            "Feature": feature_names,
            "Mean(|SHAP|)": mean_abs_shap,
            "Normalized Importance": normalized_importance,
        }
    ).sort_values(by="Mean(|SHAP|)", ascending=False).reset_index(drop=True)
    shap_table.insert(0, "Rank", np.arange(1, len(shap_table) + 1))

    slug_source = model_name or getattr(model_class, "__name__", "best_model")
    model_slug = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(slug_source)).strip("_") or "best_model"
    _save_shap_plots(shap, shap_values, explain_points, feature_names, model_slug)
    _save_shap_tables(shap_table, model_slug)

    return shap_table
def weight_analysis(model_class, data, inputs, output, num_epochs, hidden_sizes):
    X_train, X_test, y_train, y_test = read_prep_data(inputs)
    input_size = X_train.shape[1]

    output_size = y_train.shape[1] if hasattr(y_train, "shape") and len(y_train.shape) > 1 else 1
    model = make_torch_model(model_class, input_size, hidden_sizes, output_size)
    _, _, r2, model, _ = fit_model(model, X_train, X_test, y_train, y_test, num_epochs)
    print("R2:", r2)

    weight_importances = {}
    with torch.no_grad():
        for i, feature in enumerate(inputs):
            importance = np.abs(model.hidden_layers[0].weight[:, i].numpy()).mean()
            weight_importances[feature] = importance

    weight_table = pd.DataFrame(list(weight_importances.items()), columns=['Feature', 'Weight Importance'])
    return weight_table


def jackknife_sensitivity_analysis(model_class, data, inputs, output, num_epochs, hidden_sizes):
    base_model_r2, _, _, _, _, _, run, _ = repeat_fit_model(model_class,
            num_repeats, num_epochs, hidden_sizes)
    sensitivities = {}
    variances = []
    _num_repeats = input("Number of repeat [10]:")
    _num_repeats = int(_num_repeats) if _num_repeats else 10
    print("Please wait ...")

    for input_feature in inputs:
        reduced_inputs = [f for f in inputs if f != input_feature]
        reduced_r2_list = []
        for _ in range(_num_repeats):
            reduced_r2, _, _, _, _, _, _run, _ = repeat_fit_model(
                model_class, 1, num_epochs, hidden_sizes, features=reduced_inputs
            )
            if reduced_r2 is not None:
                reduced_r2_list.append(reduced_r2)
        if not reduced_r2_list or base_model_r2 is None:
            continue
        reduced_r2_mean = np.mean(reduced_r2_list)
        sensitivity = base_model_r2 - reduced_r2_mean
        variance = np.var(reduced_r2_list)
        sensitivities[input_feature] = sensitivity
        variances.append(variance)

    sensitivity_table = pd.DataFrame(list(sensitivities.items()), columns=['Feature', 'Sensitivity'])
    sensitivity_table['Variance'] = variances
    return sensitivity_table

def evaluate_feature_subset_on_training_split(
    model_class,
    features,
    num_epochs,
    hidden_sizes,
    validation_size=0.25,
):
    """Score a feature subset using only the training partition.

    The held-out test partition from prep_data is intentionally discarded here.
    Feature selection is a model-selection step, so candidate subsets must be
    chosen on an inner train/validation split of the training data and the final
    test split must remain untouched until final model evaluation.
    """
    X_train_full, _X_test_unused, y_train_full, _y_test_unused = read_prep_data(features)
    if len(X_train_full) < 3:
        return None

    inner_test_size = min(max(float(validation_size), 0.1), 0.5)
    X_fit, X_val, y_fit, y_val = train_test_split(
        X_train_full,
        y_train_full,
        test_size=inner_test_size,
        random_state=data_seed,
    )
    X_fit, X_val, _ = apply_missing_value_pipeline(X_fit, X_val, verbose=False)
    if X_fit.empty or X_val.empty:
        return None

    scores = []
    input_size = X_fit.shape[1]
    output_size = as_2d_float_array(y_fit, "y_fit").shape[1]
    for i in range(int(num_repeats)):
        try:
            if is_torch_model(model_class):
                model = make_torch_model(model_class, input_size, hidden_sizes, output_size)
                if not validate_hidden_size_compatibility(model, hidden_sizes):
                    continue
                _predictions, _mse, r2, _model, _report = fit_model(
                    model,
                    X_fit,
                    X_val,
                    y_fit,
                    y_val,
                    num_epochs,
                    run=i,
                    optimization_scope="weights",
                )
            else:
                model = model_class(model_seed + i)
                _predictions, _mse, r2, _model = fit_sklearn_model(model, X_fit, X_val, y_fit, y_val)
        except (TypeError, ValueError, RuntimeError) as err:
            print(f"Feature-selection run {i + 1} failed for {features}: {err}")
            continue
        if r2 is not None:
            scores.append(float(r2) * 100.0)

    if not scores:
        return None
    return float(np.mean(scores))


def backward_feature_elimination(model_class, data, inputs, output, num_epochs, hidden_sizes):
    mean_r2 = evaluate_feature_subset_on_training_split(model_class, inputs, num_epochs, hidden_sizes)
    if mean_r2 is None:
        print("Backward elimination skipped because the baseline model did not train successfully.")
        return pd.DataFrame(columns=["features", "R2"])
    best_r2 = mean_r2
    print("Using all features")
    print("Features:", inputs)
    print("R2:", mean_r2)
    candidates = list(inputs)
    rows = [] # Rows of a table to show the features and the result
    rows.append({"features": ", ".join(inputs), "R2": round(mean_r2,2)})
    while(True):
        if len(candidates) <= 1:
            print("Only one feature remains; stopping backward elimination.")
            break
        results = {}
        for feature in candidates:
            # Selet other features except for current feature
            features = [f for f in candidates if f != feature]
            mean_r2 = evaluate_feature_subset_on_training_split(model_class, features, num_epochs, hidden_sizes)
            if mean_r2 is None:
                continue
            results[feature] = mean_r2
            print("---------------------------------------")
            # Results of removing the feature
            print("Removing:", feature)
            print("Features:", features)
            print("R2:", mean_r2)
            rows.append({"features": ", ".join(features), "R2": round(mean_r2,2)})

        if not results:
            print("No valid feature-elimination candidates were produced; stopping.")
            break

        max_results =  max(results.values())
        max_results_key = max(results, key=results.get)

        if max_results > best_r2:
            candidates.remove(max_results_key)
            best_r2 = max_results
            print('=================================')
            print('step: ' + str(len(inputs) - len(candidates)))
            print(candidates)
            print('Best R2: ' + str(max_results))
            print('================================')
        else:
            break

    print('\n\n')
    print('=============== Final Features ==================')
    print('Selected Features: ')
    print(candidates)
    print('Final R2: ' + str(best_r2))
    print('Elminated Features: ')
    print(set(data.columns).difference(candidates))

    table = pd.DataFrame(data=rows)
    return table

# Finds the combinaiton of features that best predict y
# This method add features one by one
# Search about Forward Feature Selction
def forward_feature_selection(model_class, data, inputs, output, num_epochs, hidden_sizes):
    candidates = []
    best_r2 = -1
    rows = [] # Rows of a table to show the features and the result
    while(True):
        results = {}
        # for all features except for the candidates
        for feature in [feature for feature in inputs if feature not in candidates]:
            if len(candidates) == 0:
                features = [feature]
            else:
                features = [feature] + candidates

            mean_r2 = evaluate_feature_subset_on_training_split(model_class, features, num_epochs, hidden_sizes)
            if mean_r2 is None:
                continue
            results[feature] = mean_r2
            print("--------------------------------")
            print("Adding feature ", feature)
            print("Features:", features)
            print("R2:", mean_r2)
            rows.append({"features": ", ".join(features), "R2": round(mean_r2,2)})

        if not results:
            print("No valid feature-selection candidates were produced; stopping.")
            break

        max_results =  max(results.values())
        max_results_key = max(results, key=results.get)

        if max_results > best_r2:
            candidates.append(max_results_key)
            best_r2 = max_results
            print('=================================')
            print('step: ' + str(len(candidates)))
            print(candidates)
            print('Best R2: ' + str(max_results))
            print('================================')
        else:
            break

    print('\n\n')
    print('=============== Final Features ==================')
    print('Selected Features: ')
    print(candidates)
    print('Final R2: ' + str(best_r2))
    print('Elminated Features: ')
    print(set(data.columns).difference(candidates))

    table = pd.DataFrame(data=rows)
    return table


# Repeats an fit_model to get average of results
def repeat_fit_model(
    model_class,
    num_repeats,
    num_epochs,
    hidden_sizes,
    display_steps=False,
    features=None,
    optimization_scope="weights",
    optimize_feature_indexes=None,
    pretrained_state_dict_path=None,
):
    """Run repeated training/evaluation and summarize successful runs.

    Returns eight values in all cases: mean R² (%), std R² (%), mean MSE,
    best predictions, best R² (%), per-run R² list (%), best run index, and
    the optimized-input report from the best run.
    """
    X_train, X_test, y_train, y_test = read_prep_data(features)
    X_train, X_test, _dropped_missing_columns = apply_missing_value_pipeline(
        X_train,
        X_test,
        verbose=False,
    )
    r2_list = []
    mse_list = []
    max_r2 = -float("inf")
    max_run = 0
    best_preds = None
    best_optimized_inputs_report = None
    input_size = X_train.shape[1]
    output_size = as_2d_float_array(y_train, "y_train").shape[1]

    for i in range(int(num_repeats)):
        try:
            if is_torch_model(model_class):
                model = make_torch_model(model_class, input_size, hidden_sizes, output_size)
                if not validate_hidden_size_compatibility(model, hidden_sizes):
                    print(f"Skipping incompatible hidden sizes {hidden_sizes} for {model_class.__name__}.")
                    continue
                predictions, mse, r2, model, optimized_inputs_report = fit_model(
                    model,
                    X_train,
                    X_test,
                    y_train,
                    y_test,
                    num_epochs,
                    display_steps=display_steps,
                    run=i,
                    optimization_scope=optimization_scope,
                    optimize_feature_indexes=optimize_feature_indexes,
                    pretrained_state_dict_path=pretrained_state_dict_path,
                )
            else:
                model = model_class(model_seed + i)
                predictions, mse, r2, model = fit_sklearn_model(model, X_train, X_test, y_train, y_test)
                optimized_inputs_report = None
        except (TypeError, ValueError, RuntimeError) as err:
            print(f"Run {i + 1} failed: {err}")
            continue

        if r2 is None or mse is None or predictions is None:
            continue

        if r2 > max_r2:
            max_r2 = r2
            max_run = i
            best_preds = predictions
            best_optimized_inputs_report = optimized_inputs_report

        r2_list.append(float(r2) * 100.0)
        mse_list.append(float(mse))

    if display_steps:
        print(r2_list)

    if not r2_list:
        return None, None, None, None, None, [], max_run, None

    mean_r2 = float(np.mean(r2_list))
    mean_mse = float(np.mean(mse_list))
    std_r2 = float(np.std(r2_list))
    return (
        mean_r2,
        std_r2,
        mean_mse,
        best_preds,
        float(max_r2) * 100.0,
        r2_list,
        max_run,
        best_optimized_inputs_report,
    )


############################### Start of Program ###################
dataset_path = "convert/stclean.csv"
selected_saved_run, loaded_optimization_scope = prompt_saved_run_choice()
prepared_X_train, prepared_outputs, use_auto_feature_selection, reused_prep_data = prepare_or_reuse_data(
    dataset_path=dataset_path,
    prep_folder="prep_data",
)

saved_inputs = selected_saved_run.get("inputs") if selected_saved_run else None
saved_outputs = selected_saved_run.get("outputs") if selected_saved_run else None
run_inputs = saved_inputs if saved_inputs else prepared_X_train.columns.tolist()
run_outputs_for_validation = saved_outputs if saved_outputs else prepared_outputs
run_inputs, _ = remove_leakage_inputs_for_training(
    run_inputs,
    output_features=run_outputs_for_validation,
    context="pre_training",
)
if not run_inputs:
    print("No input columns remain after automatic target-leakage removal.")
    exit()
run_optional_future_quality_inputs = optional_input_overrides_for_selection(
    run_inputs, run_outputs_for_validation
)
X_train, X_test, y_train, y_test = read_prep_data(
    inputs=run_inputs,
    prep_folder="prep_data",
    optional_future_quality_inputs=run_optional_future_quality_inputs,
)

# After loading, get the column names from X_train
inputs = X_train.columns.tolist()
if saved_outputs:
    missing_saved_outputs = [col for col in saved_outputs if col not in y_train.columns.tolist()]
    if missing_saved_outputs:
        print(
            "Saved run output columns are missing from prep_data: "
            f"{missing_saved_outputs}. Please rebuild prep_data or choose a different saved run."
        )
        exit()
    y_train = y_train[saved_outputs]
    y_test = y_test[saved_outputs]
outputs = y_train.columns.tolist()
output = outputs
try:
    X_train, X_test, dropped_missing_columns = apply_missing_value_pipeline(X_train, X_test)
except ValueError as err:
    print(f"Missing-value pipeline failed: {err}")
    exit()
if dropped_missing_columns:
    inputs = X_train.columns.tolist()
    run_inputs = inputs

data = X_train
print_numbered_feature_list("Selected Input Features", inputs, ANSI_GREEN)
print_numbered_feature_list("Selected Output Features", outputs, ANSI_BLUE)
feature_importance_source_df = _load_feature_importance_source_dataframe(
    dataset_path,
    X_train,
    X_test,
    y_train,
    y_test,
)
generate_feature_importance_reports(feature_importance_source_df)
generate_prescriptive_optimization_report(feature_importance_source_df)
generate_model_comparison_report(X_train, X_test, y_train, y_test)
active_features = list(inputs)
if selected_saved_run:
    print("Loaded saved run metadata; reusing prepared inputs/outputs.")
    loaded_hidden_size_groups = normalize_hidden_size_groups(
        selected_saved_run.get("hidden_size_groups"),
        list_hidden_sizes,
    )
    loaded_epoch_candidates = [
        int(e)
        for e in selected_saved_run.get("epoch_candidates", [])
        if str(e).isdigit() and int(e) > 0
    ]
    loaded_repeat_count = selected_saved_run.get("repeat_count")
    if isinstance(loaded_repeat_count, str) and loaded_repeat_count.isdigit():
        loaded_repeat_count = int(loaded_repeat_count)
    if not isinstance(loaded_repeat_count, int) or loaded_repeat_count < 1:
        loaded_repeat_count = None
else:
    loaded_hidden_size_groups = [list(group) for group in list_hidden_sizes]
    loaded_epoch_candidates = []
    loaded_repeat_count = None
    ans = input("Confirm these numbered inputs/outputs from prep_data folder [y]: ").strip().lower()
    if ans not in ("", "y", "yes"):
        print("Please re-run and choose your preferred input/output features.")
        exit()

# Dynamically collect all neural-network model classes from the module
nn_models = [
    member for name, member in inspect.getmembers(models, inspect.isclass)
    if issubclass(member, models.nn.Module) and member.__module__ == models.__name__
]
sklearn_models = [
    (name, factory) for name, factory in SKLEARN_MODEL_FACTORIES.items()
]
models = nn_models + [factory for _, factory in sklearn_models]
model_names = [model.__name__ for model in nn_models] + [name for name, _ in sklearn_models]
if selected_saved_run and selected_saved_run.get("model_name") in model_names:
    selected_models = [model_names.index(selected_saved_run["model_name"])]
    print(f"Using saved model: {selected_saved_run['model_name']}")
else:
    # User input for selecting the model and number of epochs
    answer = input("\n".join([str(i) + ")" + name for i,name in enumerate(model_names)]) \
            + "\nSelect one or several models (separated with space) [all]:")

    if not answer:
        answer = "all"

    try:
        selected_model_names = parse_multi_select(answer, model_names, allow_all=True)
    except ValueError as err:
        print(f"Invalid model selection: {err}")
        exit()

    if selected_model_names is None:
        selected_models = list(range(len(models)))
    else:
        selected_models = [model_names.index(name) for name in selected_model_names]

print("Selected Models:", [model_names[i] for i in selected_models])
best_model_index = selected_models[0]
max_model_index = best_model_index
has_nn_model = any(is_torch_model(models[i]) for i in selected_models)
if has_nn_model:
    print_ann_training_robustness_notes()

default_epochs = list(list_epochs)
if loaded_epoch_candidates:
    default_epochs = loaded_epoch_candidates
answer = " ".join([str(e) for e in default_epochs])
optimization_scope_answer = ask_with_default(
    "Optimization scope for NN training [weights|inputs|both]",
    loaded_optimization_scope if loaded_optimization_scope else "weights",
)
try:
    optimization_scope = parse_optimization_scope(optimization_scope_answer)
except ValueError as err:
    print(err)
    exit()
if optimization_scope in ("inputs", "both") and not has_nn_model:
    print(
        "Input optimization is only available for neural-network models. "
        "Selected non-NN models will train/evaluate weights only."
    )
optimize_feature_indexes = None
if optimization_scope in ("inputs", "both"):
    optimize_feature_indexes = prompt_optimized_input_features(active_features)
    if optimize_feature_indexes is None:
        print("Input optimization target: all selected input features.")
    else:
        selected_names = [active_features[idx] for idx in optimize_feature_indexes]
        print("Input optimization target features:", selected_names)
change_training_options = False
if selected_saved_run:
    change_training_options = input(
        "Change training options (epochs/hidden sizes/repeats)? [n]: "
    ).strip().lower() in ("y", "yes")

if not selected_saved_run or change_training_options:
    answer = ask_with_default(
        "Enter epochs (e.g. '20 50 100', add 'cv' for cross-val early stop, or 0 to skip)",
        " ".join([str(e) for e in default_epochs]),
    )

list_epochs, use_cv_early_stop = parse_epochs_input(answer, default_epochs)
if answer != "0":
    if not list_epochs and not use_cv_early_stop:
        print("No valid epoch values were provided.")
        exit()

    if list_epochs and has_nn_model:
        print("Manual epoch candidates:", list_epochs)
    if use_cv_early_stop and has_nn_model:
        print("Cross-validation early-stop epoch search is enabled.")

    if has_nn_model and (not selected_saved_run or change_training_options):
        answer = ask_with_default(
            "Enter hidden sizes (groups split by '#', e.g. '10 5 # 15 10 3')",
            " # ".join([" ".join([str(v) for v in hs]) for hs in loaded_hidden_size_groups]),
        )
        if answer:
           list_hidden_sizes = []
           hs = answer.split("#")
           for ans in hs:
              ans = ans.strip()
              if not ans:
                  continue
              h = [int(a) for a in ans.split() if a.isdigit() and int(a) > 0]
              if h:
                  list_hidden_sizes.append(h)
        if not list_hidden_sizes:
           print("No valid hidden size values were provided.")
           exit()
    elif has_nn_model:
        list_hidden_sizes = [list(group) for group in loaded_hidden_size_groups]
        use_cv_early_stop = False
    else:
        list_hidden_sizes = [[]]
        use_cv_early_stop = False

    if not selected_saved_run or change_training_options:
        repeat_default = loaded_repeat_count if loaded_repeat_count is not None else num_repeats
        answer = ask_with_default("Enter the number of repeating predictions", repeat_default)
        if answer:
           num_repeats = int(answer)
           if num_repeats < 1:
               print("Repeat count must be at least 1.")
               exit()

    if use_auto_feature_selection:
        print("\n================= Auto Feature Selection =================")
        print("1. Backward Feature Elimination")
        print("2. Forward Feature Selection")
        auto_method = input("Select method [1]: ").strip() or "1"

        reference_model = models[selected_models[0]]
        reference_model_name = model_names[selected_models[0]]
        reference_epochs = list_epochs[0] if list_epochs else best_epochs
        reference_hidden_sizes = list_hidden_sizes[0] if is_torch_model(reference_model) else []

        print(
            f"Running feature selection with {reference_model_name}, "
            f"epochs={reference_epochs}, hidden_sizes={reference_hidden_sizes if reference_hidden_sizes else 'N/A'}"
        )

        if auto_method == "2":
            auto_table = forward_feature_selection(
                reference_model,
                data,
                list(inputs),
                output,
                reference_epochs,
                reference_hidden_sizes
            )
        else:
            auto_table = backward_feature_elimination(
                reference_model,
                data,
                list(inputs),
                output,
                reference_epochs,
                reference_hidden_sizes
            )

        active_features = infer_selected_features_from_table(auto_table, inputs)
        print(f"Auto-selected features: {active_features}")
        print("==========================================================\n")


    best_mean_r2 = -1000
    best_mse = -1000
    best_hidden_sizes = []
    best_r2 = -1000
    best_run = 0
    max_epochs = -1
    max_hidden_sizes = []
    best_epochs = -1
    results = []
    model_best_predictions = {}
    best_optimized_inputs_report = None
    # for all models
    for model_index in selected_models:
        model_class = models[model_index]
        model_name = model_names[model_index]
        is_nn = is_torch_model(model_class)
        hidden_size_candidates = list_hidden_sizes if is_nn else [[]]
        for hidden_sizes in hidden_size_candidates:
            epoch_candidates = sorted(set(list_epochs)) if is_nn else [1]
            if is_nn and use_cv_early_stop:
                cv_epoch = select_epochs_with_cv_early_stopping(
                    model_class,
                    hidden_sizes,
                    features=active_features,
                )
                if cv_epoch is not None and cv_epoch > 0:
                    epoch_candidates.append(cv_epoch)
                    print(
                        f"[{model_name} | {hidden_sizes}] CV early-stop suggested epochs: {cv_epoch}"
                    )
                else:
                    print(
                        f"[{model_name} | {hidden_sizes}] CV early-stop failed; using manual epochs only."
                    )

            epoch_candidates = sorted(set(epoch_candidates))
            for num_epochs in epoch_candidates:
                pretrained_weights_path = (
                    selected_saved_run.get("weights_path")
                    if selected_saved_run and is_nn and selected_saved_run.get("has_weights")
                    else None
                )
                # Apply model on data for N repeats and get predictions, mse and r2
                mean_r2, std_r2, mean_mse, model_best_preds, max_r2, r2_list, max_run, optimized_inputs_report = repeat_fit_model(
                    model_class,
                    num_repeats, num_epochs, hidden_sizes, display_steps=True, features=active_features, optimization_scope=optimization_scope, optimize_feature_indexes=optimize_feature_indexes, pretrained_state_dict_path=pretrained_weights_path)

                # Keep best seed to generate the same predictions later
                if mean_r2 is None:
                    continue

                cv_metrics = cross_validate_model(
                    model_class,
                    num_epochs,
                    hidden_sizes,
                    features=active_features,
                    folds=MODEL_EVALUATION_CV_FOLDS,
                    optimization_scope=optimization_scope if is_nn else "weights",
                    optimize_feature_indexes=optimize_feature_indexes if is_nn else None,
                )

                if max_r2 > best_r2:
                    best_r2 = max_r2
                    model_best_predictions[model_name] = model_best_preds
                    best_optimized_inputs_report = optimized_inputs_report
                    max_model_index = model_index
                    max_epochs = num_epochs
                    max_hidden_sizes = hidden_sizes
                    best_run = max_run

                if mean_r2 > best_mean_r2:
                    best_mean_r2 = mean_r2
                    best_mse = mean_mse
                    best_model_index = model_index
                    best_epochs = num_epochs
                    best_hidden_sizes = hidden_sizes

                total_nodes = sum(hidden_sizes) if hidden_sizes else 0

                mean_rmse = float(np.sqrt(mean_mse)) if mean_mse is not None else np.nan
                result = {
                        "model":model_name,
                        "R2": round(mean_r2,1),
                        "MSE": round(mean_mse,2),
                        "R2 std": round(std_r2, 1),
                        "R2 List": [round(x, 1) for x in r2_list],
                        "Mean R²": round(cv_metrics["mean_r2"], 3) if cv_metrics else np.nan,
                        "Std R²": round(cv_metrics["std_r2"], 3) if cv_metrics else np.nan,
                        "Mean RMSE": round(cv_metrics["mean_rmse"], 3) if cv_metrics else round(mean_rmse, 3),
                        "Std RMSE": round(cv_metrics["std_rmse"], 3) if cv_metrics else np.nan,
                        "CV folds": cv_metrics["folds"] if cv_metrics else 0,
                        "hidden sizes": hidden_sizes if hidden_sizes else "N/A",
                        "total hs": total_nodes,
                        "epochs": num_epochs if is_nn else "N/A",
                        }
                results.append(result)

    # Create a table for results. Some model/config combinations can fail and return no
    # valid R2 values, so guard against missing/empty result sets.
    expected_result_cols = [
        "model", "R2", "MSE", "R2 std", "R2 List",
        "Mean R²", "Std R²", "Mean RMSE", "Std RMSE", "CV folds",
        "hidden sizes", "total hs", "epochs"
    ]
    results_table = pd.DataFrame(data=results)
    if results_table.empty:
        results_table = pd.DataFrame(columns=expected_result_cols)
    else:
        missing_cols = [col for col in expected_result_cols if col not in results_table.columns]
        for col in missing_cols:
            results_table[col] = np.nan
        # Sort methods by R2 only when the column exists and contains data.
        if results_table["R2"].notna().any():
            results_table = results_table.sort_values(by="R2", ascending=False)
        else:
            print("No valid R2 values were produced for the selected model/config combinations.")

    if results_table["R2"].isna().all():
        print("Training finished, but no successful runs produced R2 values. Skipping reporting/plots.")
        results_table.to_csv("exp.csv", index=False)
        exit()

    latex_table = results_table.copy()
    # Create and save latex code for table
    latex_table["R2"] = latex_table.apply(lambda row: f"{row['R2']} ± {row['R2 std']}", axis=1)
    latex_table = latex_table.drop(columns=["R2 List","R2 std"])
    results_table_latex = generate_latex_table(latex_table,
            caption="Results of different models", label="models")
    with open(os.path.join("tables", "results.tex"), 'w', encoding='utf-8') as f:
        print(results_table_latex, file=f)

    # Plot the performance of models across different parameters
    print("Generating plots ...")
    plot_model_performance(results_table)

    best_model = models[best_model_index]
    best_model_name = model_names[best_model_index]
    # Show results
    max_model_name = model_names[max_model_index]

    print("============ Results for models =========================")
    print(results_table)
    print("========================== Best Mean Model ===============================")
    print("Best Mean R-Squred:", best_mean_r2)
    print("Best model with better mean R-Squred:", best_model_name)
    print("Best Hidden sizes:", best_hidden_sizes)
    print("Best epochs:", best_epochs)
    print("=========================== Max Model ================================")
    print("Best model with better max R-Squred:", max_model_name)
    print("Best Hidden sizes:", max_hidden_sizes)
    print("Best epochs:", max_epochs)
    print("Max R-Squred:", best_r2)

    results_table.to_csv("exp.csv")

    X_train, X_test, y_train, y_test = read_prep_data(active_features)
    generate_target_shap_analysis(X_train, X_test, y_train, y_test, targets=outputs)

    # Show and save the plot for best results
    best_predictions = model_best_predictions.get(max_model_name)
    if best_predictions is None:
        print("Best model did not produce predictions. Skipping reporting/plots.")
        exit()
    output_title = ", ".join(outputs)
    title = "Prediction of " + output_title + " with " + max_model_name + " epochs:" + str(max_epochs)
    file_name = f"R2-{best_r2:.2f}-" + max_model_name + "-" + "-".join(outputs) + ".png"

    y_test_values = y_test.values if hasattr(y_test, "values") else y_test
    best_model_metrics = compute_regression_report_metrics(y_test_values, best_predictions)
    if best_model_metrics:
        print("===================== Best Selected Model Report =====================")
        print(f"Selected model name                    {max_model_name}")
        print(f"R-Squared                              {best_r2/100:.4f}")
        print(
            f"Correlation coefficient                {best_model_metrics['Correlation coefficient']:.4f}"
        )
        print(
            f"Mean absolute error                    {best_model_metrics['Mean absolute error']:.4f}"
        )
        print(
            f"Root mean squared error                {best_model_metrics['Root mean squared error']:.4f}"
        )
        print(
            f"Relative absolute error                {best_model_metrics['Relative absolute error']:.4f} %"
        )
        print(
            f"Root relative squared error            {best_model_metrics['Root relative squared error']:.4f} %"
        )
        print(
            f"Total Number of Instances              {best_model_metrics['Total Number of Instances']}"
        )
        print("=====================================================================")

    if best_optimized_inputs_report is not None:
        print("\n================= Optimized Input Report =================")
        selected_names = (
            active_features
            if optimize_feature_indexes is None
            else [active_features[idx] for idx in optimize_feature_indexes]
        )
        mean_abs_delta = best_optimized_inputs_report["mean_abs_delta_per_feature"]
        report_df = pd.DataFrame(
            {
                "feature": active_features,
                "mean_abs_delta_normalized": mean_abs_delta,
            }
        ).sort_values(by="mean_abs_delta_normalized", ascending=False)
        report_df["optimized"] = report_df["feature"].isin(selected_names)
        report_path = os.path.join("tables", "optimized-inputs-summary.csv")
        report_df.to_csv(report_path, index=False)
        print(f"Saved summary: {report_path}")
        print(report_df)

        original_df = pd.DataFrame(
            best_optimized_inputs_report["original_inputs"], columns=active_features
        )
        optimized_df = pd.DataFrame(
            best_optimized_inputs_report["optimized_inputs"], columns=active_features
        )
        original_df.to_csv(os.path.join("tables", "optimized-inputs-before.csv"), index=False)
        optimized_df.to_csv(os.path.join("tables", "optimized-inputs-after.csv"), index=False)

        fig, ax = plt.subplots(figsize=(12, 6))
        ax.bar(report_df["feature"], report_df["mean_abs_delta_normalized"])
        ax.set_title("Mean absolute input change after input optimization")
        ax.set_xlabel("Input feature")
        ax.set_ylabel("Mean abs change (normalized)")
        ax.tick_params(axis="x", rotation=45)
        ax.grid(True, axis="y", alpha=0.3)
        fig.tight_layout()
        fig.savefig(os.path.join("plots", "optimized-inputs-change-summary.png"), format="png")
        plt.close(fig)
        print("Saved plot: plots/optimized-inputs-change-summary.png")
        print("==========================================================")

    print("\n\n")
    print("Plot was saved in plots folder")
    answer = input("Do you want to see them? [y]:")
    if len(outputs) == 1:
        target_series = y_test[outputs[0]]
        pred_series = best_predictions[:, 0] if best_predictions.ndim > 1 else best_predictions
        if answer == "y" or answer == "yes":
            plot_results(pred_series, target_series, title, file_name, show_plot=True)
        else:
            plot_results(pred_series, target_series, title, file_name, show_plot=False)
    else:
        print("Skipping scatter plot for multi-output predictions.")

    # Save results of predicitons in a file named results.csv
    results_df = pd.DataFrame()
    for i, output_name in enumerate(outputs):
        results_df[f"{output_name}_actual"] = y_test[output_name].values
        pred_col = best_predictions[:, i] if best_predictions.ndim > 1 else best_predictions
        results_df[f"{output_name}_predictions"] = np.round(pred_col, 2)
    results_df.to_csv("results.csv", index=False)
    print("Predictions of best model were saved in results.csv")
    weights_path = None
    if is_torch_model(models[max_model_index]):
        os.makedirs(CHECKPOINTS_DIR, exist_ok=True)
        checkpoint_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", f"{max_model_name}_R2-{best_r2:.2f}_run-{best_run}.pt")
        weights_path = os.path.join(CHECKPOINTS_DIR, checkpoint_name)
        input_size = X_train.shape[1]
        output_size = as_2d_float_array(y_train, "y_train").shape[1]
        best_model_instance = make_torch_model(models[max_model_index], input_size, max_hidden_sizes, output_size)
        checkpoint_predictions, checkpoint_mse, checkpoint_r2, best_model_instance, _ = fit_model(
            best_model_instance,
            X_train,
            X_test,
            y_train,
            y_test,
            max_epochs,
            display_steps=False,
            run=best_run,
            optimization_scope="weights",
            optimize_feature_indexes=None,
            pretrained_state_dict_path=None,
        )
        checkpoint_ready = (
            checkpoint_r2 is not None
            and checkpoint_predictions is not None
            and getattr(best_model_instance, "input_scaler_", None) is not None
            and getattr(best_model_instance, "target_scaler_", None) is not None
        )
        if checkpoint_ready:
            torch.save(best_model_instance.state_dict(), weights_path)
            print(f"Saved model checkpoint: {weights_path}")

            try:
                operator_controls = tuple(
                    variable for variable in OPERATOR_RECOMMENDATION_VARIABLES
                    if variable in active_features
                )
                if not operator_controls:
                    raise RecommendationError(
                        "None of the requested operator recommendation settings are present "
                        "in the trained input features."
                    )
                recommendation_grid_steps = 3 if len(operator_controls) > 4 else 7
                recommendation = recommend_operating_conditions(
                    current_conditions=X_test.iloc[0],
                    trained_model=best_model_instance,
                    input_features=active_features,
                    output_features=outputs,
                    historical_inputs=X_train,
                    historical_targets=y_train,
                    control_variables=operator_controls,
                    grid_steps=recommendation_grid_steps,
                )
                recommendation_path = os.path.join("tables", "recommended-operating-conditions.json")
                with open(recommendation_path, "w", encoding="utf-8") as fp:
                    json.dump(recommendation, fp, indent=2, ensure_ascii=False, default=float)
                operator_report_path = save_operator_report(recommendation)
                print_industrial_operator_demo(
                    recommendation=recommendation,
                    input_features=active_features,
                    output_features=outputs,
                )
                print(f"Saved recommendation: {recommendation_path}")
                print(f"Saved operator workbook: {operator_report_path}")
            except (RecommendationError, ValueError, KeyError, TypeError) as err:
                print(f"Recommendation engine skipped: {err}")
        else:
            print("Checkpoint save skipped because the best model could not be retrained safely.")
            weights_path = None

    save_run_summary(
        model_name=max_model_name,
        inputs=active_features,
        outputs=outputs,
        best_r2=best_r2,
        epoch_candidates=list_epochs,
        hidden_size_groups=list_hidden_sizes if is_torch_model(models[max_model_index]) else [],
        repeat_count=num_repeats,
        optimization_scope=optimization_scope,
        weights_path=weights_path,
    )
    answer = input("Do you want to see them? [y]:")
    if answer == "y" or answer == "yes":
       print("======= Predictions of best model:", best_model_name)
       print(results_df)

best_model = models[best_model_index]
best_model_name = model_names[best_model_index]
if is_torch_model(best_model):
    while True:
        print("\n\n")
        print(f"================= Feature Selection ({best_model_name}:{best_epochs} epochs, {best_hidden_sizes}) ======")
        print("\nPlease select a feature selection or sensitivity analysis method:\n")
        print("1. Backward Feature Elimination")
        print("2. Forward Feature Selection")
        print("3. Weight Analysis")
        print("4. Jackknife Sensitivity Analysis (Node Deletion Sensitivity)")
        print("5. SHAP (SHapley Additive exPlanations)")
        print("q. Quit")

        answer = input("Enter the number of the method you want to run (or 'q' to quit): ").strip().lower()

        if answer == '1':
            print("============================= Backward Feature Elimination =============")
            backward_table = backward_feature_elimination(best_model, data, inputs, output, best_epochs, best_hidden_sizes)
            print("------------ backward feature elimination ---------------")
            print(backward_table)

            backward_table_latex = generate_latex_table(backward_table, caption="Results of Backward Feature Elimination", label="backward")
            with open(os.path.join("tables", "backward.tex"), 'w', encoding='utf-8') as f:
                print(backward_table_latex, file=f)

        elif answer == '2':
            print("============================= Forward Feature Selection ================")
            forward_table = forward_feature_selection(best_model, data, inputs, output,
                                                      best_epochs, best_hidden_sizes)
            print("\n")
            print("------------ forward feature selection ---------------")
            print(forward_table)
            forward_table_latex = generate_latex_table(forward_table, caption="Results of Forward Feature Selection for different features", label="forward")
            with open(os.path.join("tables", "forward.tex"), 'w', encoding='utf-8') as f:
                print(forward_table_latex, file=f)

        elif answer == '3':
            print("============================= Weight Analysis =============")
            weight_table = weight_analysis(best_model, data, inputs, output, best_epochs, best_hidden_sizes)
            print("------------ weight analysis ---------------")
            weight_table = weight_table.sort_values(by='Weight Importance',
                    ascending=False)
            print("Most important features:")
            print(weight_table)
            weight_table_latex = generate_latex_table(weight_table, caption="Results of Weight Analysis", label="weight_analysis")
            with open(os.path.join("tables", "weight-analysis.tex"), 'w', encoding='utf-8') as f:
                print(weight_table_latex, file=f)

        elif answer == '4':
            print("============================= Jackknife Sensitivity Analysis =============")
            jackknife_table = jackknife_sensitivity_analysis(best_model,
                    data, inputs, output, best_epochs, best_hidden_sizes)
            print("------------ jackknife sensitivity analysis ---------------")
            jackknife_table = jackknife_table.sort_values(by='Sensitivity', ascending=False)
            print(jackknife_table)
            jackknife_table_latex = generate_latex_table(jackknife_table, caption="Results of Jackknife Sensitivity Analysis", label="jackknife")
            with open(os.path.join("tables", "jackknife.tex"), 'w', encoding='utf-8') as f:
                print(jackknife_table_latex, file=f)

            important_features = jackknife_table.head(5)['Feature'].tolist()
            print("\nTop 5 important features to focus on:")
            print(important_features)

            negative_sensitivity_features = jackknife_table[jackknife_table['Sensitivity'] < 0]
            print("\nFeatures with negative sensitivity (potentially redundant or harmful):")
            print(negative_sensitivity_features)

        elif answer == '5':
            print("============================= SHAP Feature Importance =============")
            shap_table = shap_feature_importance(best_model, inputs, best_epochs, best_hidden_sizes, model_name=best_model_name)
            if shap_table is not None:
                print("------------ SHAP feature importance ---------------")
                print(shap_table)
                shap_table_latex = generate_latex_table(shap_table, caption="Results of SHAP Feature Importance", label="shap")
                with open(os.path.join("tables", "shap.tex"), 'w', encoding='utf-8') as f:
                    print(shap_table_latex, file=f)
                shap_table.to_csv(os.path.join("tables", "shap.csv"), index=False)

        elif answer == 'q':
            print("Exiting the feature selection and sensitivity analysis loop. Goodbye!")
            break
        else:
            print("Invalid choice, please try again.")
        print("\n----------------------------------------------------------")
        input("Press any key to return to main menu ...")
else:
    while True:
        print("\n\n")
        print(f"================= SHAP Analysis ({best_model_name}) ======")
        print("1. SHAP (SHapley Additive exPlanations)")
        print("q. Quit")

        answer = input("Enter the number of the method you want to run (or 'q' to quit): ").strip().lower()

        if answer == '1':
            print("============================= SHAP Feature Importance =============")
            shap_table = shap_feature_importance(best_model, inputs, best_epochs, best_hidden_sizes, model_name=best_model_name)
            if shap_table is not None:
                print("------------ SHAP feature importance ---------------")
                print(shap_table)
                shap_table_latex = generate_latex_table(shap_table, caption="Results of SHAP Feature Importance", label="shap")
                with open(os.path.join("tables", "shap.tex"), 'w', encoding='utf-8') as f:
                    print(shap_table_latex, file=f)
                shap_table.to_csv(os.path.join("tables", "shap.csv"), index=False)
        elif answer == 'q':
            print("Exiting the SHAP analysis loop. Goodbye!")
            break
        else:
            print("Invalid choice, please try again.")
        print("\n----------------------------------------------------------")
        input("Press any key to return to main menu ...")



print("----------------------Important! READ -------------------")
print("latex code for tables are saved in tables folder")
print("predictions are saved in results.csv file")
input("Plots have been saved in the 'plots' folder. (press any key to exit)")

# Visualize the model and save it on mlp_structure image
# dummy_input = torch.randn(1, input_size)
# from torchviz import make_dot
# dot = make_dot(model(dummy_input), params=dict(model.named_parameters()))
# dot.render("mlp_structure", format="png", cleanup=True)
